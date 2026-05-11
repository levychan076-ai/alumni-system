import pymysql
import pymysql.cursors
import re
import csv
import io
import os
import uuid
import anthropic
from werkzeug.utils import secure_filename
from io import BytesIO
from datetime import date, datetime
import time
from functools import wraps
from reportlab.platypus import PageBreak
from flask import flash
from flask import Flask, render_template, request, redirect, session, send_file, jsonify, url_for
from flask_mail import Mail, Message
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import pandas as pd
from werkzeug.utils import secure_filename
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
app = Flask(__name__)
app.secret_key = "secret123"



def get_db():

    railway_host = os.getenv("MYSQLHOST")
    railway_port = os.getenv("MYSQLPORT")

    # Railway Production
    if railway_host and railway_port:

        return pymysql.connect(
            host=railway_host,
            user=os.getenv("MYSQLUSER"),
            password=os.getenv("MYSQLPASSWORD"),
            database=os.getenv("MYSQLDATABASE") or os.getenv("MYSQL_DATABASE"),
            port=int(railway_port),
            cursorclass=pymysql.cursors.DictCursor,
            autocommit=False
        )

    # Local XAMPP
    return pymysql.connect(
        host="localhost",
        user="root",
        password="",
        database="alumni_system",
        port=3306,
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=False
    )

def test_database():
    """Test database connection and user_login table"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Test connection
        cursor.execute("SELECT 1 as test")
        result = cursor.fetchone()
        print(f"DEBUG - Database connection test: {result}")
        
        # Test table structure
        cursor.execute("DESCRIBE user_login")
        columns = cursor.fetchall()
        print(f"DEBUG - user_login table structure:")
        for col in columns:
            print(f"  {col}")
        
        # Test sample data
        cursor.execute("SELECT * FROM user_login LIMIT 5")
        users = cursor.fetchall()
        print(f"DEBUG - Sample user_login data:")
        for user in users:
            print(f"  {user}")
        
        cursor.close()
        db.close()
        return True
    except Exception as e:
        import traceback
        print("ERROR in get_db():", str(e))
        traceback.print_exc()
        return None

from functools import wraps
from flask import redirect, session, request

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        # Check if user is logged in (either admin/coordinator or alumni)
        if "username" not in session and "alumni_email" not in session:
            return redirect("/")
        
        # Additional validation for admin/coordinator users
        if "username" in session and session.get("user_type") not in ["Admin", "Alumni Coordinator", "Alumni"]:
            return redirect("/")
        
        return f(*args, **kwargs)
    return wrapper


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "username" not in session:
            return redirect("/")
        if session.get("user_type") != "Admin":
            return "Access Denied: Admins only", 403
        return f(*args, **kwargs)
    return wrapper


def coordinator_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "username" not in session:
            return redirect("/")
        if session.get("user_type") != "Alumni Coordinator":
            return "Access Denied: Alumni Coordinators only", 403
        return f(*args, **kwargs)
    return wrapper


def draw_layout(canvas, doc):
    width, height = LETTER


    canvas.setLineWidth(1.2)
    canvas.rect(25, 25, width - 50, height - 50)


def build_alumni_page(elements, r):

    def safe(val):
        if val is None:
            return ""
        if hasattr(val, "strftime"):
            return val.strftime("%Y-%m-%d")
        return str(val)

    # ================= STYLES (FIX: avoid re-creation issues) =================
    normal = ParagraphStyle(name="normal", fontName="Times-Roman", fontSize=10, leading=12)
    bold = ParagraphStyle(name="bold", fontName="Times-Bold", fontSize=10)
    header = ParagraphStyle(name="header", fontName="Times-Bold", fontSize=11)

    # ================= LOGO SAFE LOAD =================
    try:
        logo = Image("static/images/PNG (transparent background).png", 50, 50)
    except:
        logo = None

    if not logo:
        logo = ""

    # ================= HEADER =================
    header_tbl = Table([
        [
            logo,
            Paragraph(
                "Nueva Ecija University of Science and Technology<br/>"
                "Talavera Academic Extension Campus<br/><br/>"
                "<b>Alumni Profile</b>",
                header
            )
        ]
    ], colWidths=[60, 420])

    header_tbl.setStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ])

    elements.append(header_tbl)
    elements.append(Spacer(1, 10))

    # ================= ALUMNI INFO =================
    alumni_info = Table([
        [Paragraph("<b>Alumni Information</b>", bold), ""],

        ["Name:", f"{safe(r[2])}, {safe(r[3])} {safe(r[4])}"],
        ["Address:", safe(r[5])],
        ["Email:", safe(r[6])],
        ["Contact No.:", safe(r[7])],
    ], colWidths=[150, 330])

    alumni_info.setStyle([
        ("BOX", (0, 0), (-1, -1), 1, colors.black),
        ("SPAN", (0, 0), (1, 0)),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ])

    elements.append(alumni_info)
    elements.append(Spacer(1, 10))

    # ================= DEGREE INFO =================
    degree = Table([
        [Paragraph("<b>Degree Program</b>", bold), ""],

        ["Program:", safe(r[10])],
        ["Major:", safe(r[11])],
        ["Date of Admission:", safe(r[12])],
        ["Graduation Date:", safe(r[13])],
    ], colWidths=[150, 330])

    degree.setStyle([
        ("BOX", (0, 0), (-1, -1), 1, colors.black),
        ("SPAN", (0, 0), (1, 0)),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ])

    elements.append(degree)
    elements.append(Spacer(1, 10))

    # ================= EMPLOYMENT INFO =================
    employment = Table([
        [Paragraph("<b>Employment Information</b>", bold), ""],

        ["Status:", safe(r[15])],
        ["Job Title:", safe(r[14])],
        ["Sector:", safe(r[16])],
    ], colWidths=[150, 330])

    employment.setStyle([
        ("BOX", (0, 0), (-1, -1), 1, colors.black),
        ("SPAN", (0, 0), (1, 0)),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ])

    elements.append(employment)

# ================= para sa email =================
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USERNAME'] = 'neust.talavera.offcampus.neust.mgt@gmail.com'
app.config['MAIL_PASSWORD'] = 'cshoojrdzdtwzlzl'
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_DEFAULT_SENDER'] = ('NEUST TALAVERA OFF-CAMPUS', app.config['MAIL_USERNAME'])

mail = Mail(app)

UPLOAD_FOLDER = os.path.join("static", "uploads")
ALLOWED_PHOTO = {"jpg", "jpeg", "png", "gif", "webp"}
MAX_PHOTO_BYTES = 5 * 1024 * 1024  # 5 MB

os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def allowed_photo(filename):
    return (
            "." in filename
            and filename.rsplit(".", 1)[1].lower() in ALLOWED_PHOTO
    )


def save_photo(file_obj):
    """Save uploaded photo, return saved filename or None."""
    if not file_obj or file_obj.filename == "":
        return None
    if not allowed_photo(file_obj.filename):
        return None
    ext = file_obj.filename.rsplit(".", 1)[1].lower()
    filename = f"{uuid.uuid4().hex}.{ext}"
    file_obj.save(os.path.join(UPLOAD_FOLDER, filename))
    return filename


def delete_photo(filename):
    """Delete a photo file from disk if it exists."""
    if not filename:
        return
    path = os.path.join(UPLOAD_FOLDER, filename)
    try:
        if os.path.exists(path):
            os.remove(path)
    except Exception as e:
        print("Photo delete error:", e)


@app.route("/api/scan-relevance", methods=["POST"])
@login_required
def scan_relevance():
    data = request.get_json(force=True)
    job_title = (data.get("job_title") or "").strip()
    program = (data.get("program") or "").strip()
    major = (data.get("major") or "").strip()

    if not job_title:
        return jsonify({"relevance": "Not Related"})

    prompt = (
        f"A college graduate completed a {program} degree "
        f"with a major in {major}. "
        f"Their current job title is: '{job_title}'.\n\n"
        f"Based on this, classify how relevant their degree is to their work. "
        f"Reply with EXACTLY one of these four options and nothing else:\n"
        f"Directly Related\n"
        f"Moderately Related\n"
        f"Slightly Related\n"
        f"Not Related"
    )

    try:
        client = anthropic.Anthropic()
        message = client.messages.create(
            model="claude-opus-4-6",
            max_tokens=20,
            messages=[{"role": "user", "content": prompt}]
        )
        result = message.content[0].text.strip()

        valid = ["Directly Related", "Moderately Related", "Slightly Related", "Not Related"]
        if result not in valid:
            # Fuzzy match
            result_lower = result.lower()
            if "directly" in result_lower:
                result = "Directly Related"
            elif "moderate" in result_lower:
                result = "Moderately Related"
            elif "slightly" in result_lower:
                result = "Slightly Related"
            else:
                result = "Not Related"

        return jsonify({"relevance": result})

    except Exception as e:
        print("AI scan error:", e)
        return jsonify({"relevance": "Not Related", "error": str(e)}), 200




# ================= Admin/Coordinator Login Page =================
@app.route("/test-login")
def test_login():
    """Test route to verify database and login logic"""
    print("=== TESTING LOGIN SYSTEM ===")
    
    # Test database
    if not test_database():
        return "Database test failed"
    
    # Test specific login query
    db = get_db()
    cursor = db.cursor()
    
    try:
        # Test with sample credentials
        test_username = "admin1"
        test_password = "admin123"
        
        cursor.execute("""
            SELECT username, user_type
            FROM user_login
            WHERE username = %s AND password = %s
            LIMIT 1
        """, (test_username, test_password))
        
        user = cursor.fetchone()
        print(f"DEBUG - Test query result: {user}")
        
        if user:
            return f"Found user: {user['username']} with type: {user['user_type']}"
        else:
            return f"No user found with username='{test_username}' and password='{test_password}'"
            
    except Exception as e:
        return f"Query error: {e}"
    finally:
        cursor.close()
        db.close()

@app.route("/login-admin", methods=["GET", "POST"])
def login_admin():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        login_role = request.form.get("login_role", "admin")  # "admin" or "coordinator"

        # DEBUG: Print form values
        print(f"DEBUG - Login attempt:")
        print(f"  Username: '{username}'")
        print(f"  Password: '{password}'")
        print(f"  Role: '{login_role}'")

        # Validate required fields
        if not username or not password:
            return render_template("login_admin.html", error="Username and password are required.")

        db = get_db()
        cursor = db.cursor()

        try:
            # Query user with exact username and password match
            cursor.execute("""
                SELECT username, user_type
                FROM user_login
                WHERE username = %s AND password = %s
                LIMIT 1
            """, (username, password))

            user = cursor.fetchone()
            print(f"DEBUG - Database result: {user}")

            if user:
                db_user_type = user["user_type"]  # Database stores as 'Admin' or 'Alumni Coordinator'
                print(f"DEBUG - User found with type: '{db_user_type}'")

                # STRICT ROLE VALIDATION - Use exact case from database
                if login_role == "admin":
                    if db_user_type != "Admin":
                        print(f"DEBUG - Role mismatch: expected 'Admin', got '{db_user_type}'")
                        return render_template("login_admin.html", error="Access Denied: This account is not an Admin account.")
                    
                    session["username"] = user["username"]
                    session["user_type"] = "Admin"
                    log_activity("Admin logged in")
                    return redirect("/dashboard")
                
                elif login_role == "coordinator":
                    if db_user_type != "Alumni Coordinator":
                        print(f"DEBUG - Role mismatch: expected 'Alumni Coordinator', got '{db_user_type}'")
                        return render_template("login_admin.html", error="Access Denied: This account is not an Alumni Coordinator account.")
                    
                    session["username"] = user["username"]
                    session["user_type"] = "Alumni Coordinator"
                    log_activity("Alumni Coordinator logged in")
                    return redirect("/dashboard")

            print(f"DEBUG - No user found or invalid credentials")
            return render_template("login_admin.html", error="Invalid username or password.")

        except Exception as e:
            import traceback
            print("ERROR - Login failed:", str(e))
            traceback.print_exc()
            return render_template("login_admin.html", error="Database error. Please try again.")
        finally:
            cursor.close()
            db.close()

    return render_template("login_admin.html")

# ================= Alumni Login Page =================
@app.route("/login-alumni", methods=["GET", "POST"])
def login_alumni():
    if request.method == "POST":
        alumni_email = request.form.get("alumni_email", "").strip()
        alumni_password = request.form.get("alumni_password", "")

        print(f"DEBUG - Alumni Login Attempt:")
        print(f"  Email: '{alumni_email}'")
        print(f"  Password: '{alumni_password}'")

        db = get_db()
        cursor = db.cursor()

        try:
            cursor.execute("""
                SELECT alumni_id, last_name, first_name, middle_name, email
                FROM alumni_table
                WHERE email = %s AND alumni_password = %s
                LIMIT 1
            """, (alumni_email, alumni_password))

            alumni_user = cursor.fetchone()
            print(f"DEBUG - Database query result: {alumni_user}")
            
        except Exception as e:
            import traceback
            print("ERROR - Alumni login failed:", str(e))
            traceback.print_exc()
            alumni_user = None
        finally:
            cursor.close()
            db.close()

        if alumni_user:
            print(f"DEBUG - Login successful, setting session...")
            full_name = f"{alumni_user['first_name']} {alumni_user['last_name']}"
            session["username"] = alumni_user["email"]
            session["alumni_email"] = alumni_user["email"]  # Add this for consistency
            session["user_type"] = "Alumni"
            session["alumni_id"] = alumni_user["alumni_id"]
            session["alumni_fullname"] = full_name
            
            print(f"DEBUG - Session set: {dict(session)}")
            print(f"DEBUG - Redirecting to /my-profile")

            log_activity("Alumni logged in")
            return redirect("/my-profile")

        print(f"DEBUG - Login failed, returning to login page")
        return render_template("login_alumni.html", error="Invalid email or password.")

    return render_template("login_alumni.html")

# ================= Original Login (for backward compatibility) =================
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        # Handle admin/coordinator login
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        login_role = request.form.get("login_role", "admin")  # "admin" or "coordinator"

        # DEBUG: Print form values
        print(f"DEBUG - Original Login attempt:")
        print(f"  Username: '{username}'")
        print(f"  Password: '{password}'")
        print(f"  Role: '{login_role}'")

        # Validate required fields
        if not username or not password:
            return render_template("login.html", error="Username and password are required.")

        db = get_db()
        cursor = db.cursor()

        try:
            # Query user with exact username and password match
            cursor.execute("""
                SELECT username, user_type
                FROM user_login
                WHERE username = %s AND password = %s
                LIMIT 1
            """, (username, password))

            user = cursor.fetchone()
            print(f"DEBUG - Database result: {user}")

            if user:
                db_user_type = user["user_type"]  # Database stores as 'Admin' or 'Alumni Coordinator'
                print(f"DEBUG - User found with type: '{db_user_type}'")

                # STRICT ROLE VALIDATION - Use exact case from database
                if login_role == "admin":
                    if db_user_type != "Admin":
                        print(f"DEBUG - Role mismatch: expected 'Admin', got '{db_user_type}'")
                        return render_template("login.html", error="Access Denied: This account is not an Admin account.")
                    
                    session["username"] = user["username"]
                    session["user_type"] = "Admin"
                    log_activity("Admin logged in")
                    return redirect("/dashboard")
                
                elif login_role == "coordinator":
                    if db_user_type != "Alumni Coordinator":
                        print(f"DEBUG - Role mismatch: expected 'Alumni Coordinator', got '{db_user_type}'")
                        return render_template("login.html", error="Access Denied: This account is not an Alumni Coordinator account.")
                    
                    session["username"] = user["username"]
                    session["user_type"] = "Alumni Coordinator"
                    log_activity("Alumni Coordinator logged in")
                    return redirect("/dashboard")

            print(f"DEBUG - No user found or invalid credentials")
            return render_template("login.html", error="Invalid username or password.")

        except Exception as e:
            import traceback
            print("ERROR - Original login failed:", str(e))
            traceback.print_exc()
            return render_template("login.html", error="Database error. Please try again.")
        finally:
            cursor.close()
            db.close()

    return render_template("login.html")

# ================= para sa dashboard =================
@app.route("/dashboard")
@login_required
def dashboard():
    user_type = session.get("user_type")

    # ALUMNI users go to their own profile/dashboard view
    if user_type == "Alumni":
        return redirect("/my-profile")

    # Get dashboard data based on user role
    db = get_db()
    cursor = db.cursor()
    
    try:
        # Get comprehensive statistics for dashboard
        # Total alumni count
        cursor.execute("SELECT COUNT(*) as total FROM alumni_table")
        result = cursor.fetchone()
        total_alumni = result['total'] if result else 0
        
        # Program-specific counts
        cursor.execute("""
            SELECT 
                d.program,
                COUNT(DISTINCT a.alumni_id) as total
            FROM alumni_table a
            LEFT JOIN alumni_degree d ON a.alumni_id = d.alumni_id
            WHERE d.program IS NOT NULL
            GROUP BY d.program
        """)
        program_counts = cursor.fetchall()
        
        # Initialize program counters
        bsit_count = 0
        bsba_count = 0
        beed_count = 0
        
        for program in program_counts:
            if program['program'] == 'BSIT':
                bsit_count = program['total']
            elif program['program'] == 'BSBA':
                bsba_count = program['total']
            elif program['program'] == 'BEED':
                beed_count = program['total']
        
        # Employment statistics
        cursor.execute("""
            SELECT 
                e.employment_status,
                COUNT(DISTINCT a.alumni_id) as total
            FROM alumni_table a
            LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
            WHERE e.employment_status IS NOT NULL
            GROUP BY e.employment_status
        """)
        employment_stats = cursor.fetchall()
        
        employed = 0
        unemployed = 0
        self_employed = 0
        
        for emp in employment_stats:
            if emp['employment_status'] == 'Employed':
                employed = emp['total']
            elif emp['employment_status'] == 'Unemployed':
                unemployed = emp['total']
            elif emp['employment_status'] == 'Self-Employed':
                self_employed = emp['total']
        
        # Work relevance statistics
        cursor.execute("""
            SELECT 
                e.degree_relevance_to_work,
                COUNT(DISTINCT a.alumni_id) as total
            FROM alumni_table a
            LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
            WHERE e.degree_relevance_to_work IS NOT NULL
              AND e.employment_status != 'Unemployed'
            GROUP BY e.degree_relevance_to_work
        """)
        relevance_stats = cursor.fetchall()
        
        directly_related = 0
        for rel in relevance_stats:
            if rel['degree_relevance_to_work'] == 'Directly Related':
                directly_related = rel['total']
        
        # Get major-specific counts for each program
        cursor.execute("""
            SELECT 
                d.program,
                d.major,
                COUNT(DISTINCT a.alumni_id) as total
            FROM alumni_table a
            LEFT JOIN alumni_degree d ON a.alumni_id = d.alumni_id
            WHERE d.program IS NOT NULL AND d.major IS NOT NULL
            GROUP BY d.program, d.major
            ORDER BY d.program, d.major
        """)
        major_counts = cursor.fetchall()
        
        # Get recent activities
        cursor.execute("""
            SELECT username, user_type, activity, date_time 
            FROM activity_logs 
            ORDER BY date_time DESC 
            LIMIT 10
        """)
        recent_activities = cursor.fetchall()
        
        # Get pending requests (for admin)
        cursor.execute("SELECT COUNT(*) as pending FROM alumni_notifications WHERE status = 'pending'")
        result = cursor.fetchone()
        pending_requests = result['pending'] if result else 0
        
    except Exception as e:
        import traceback
        print("ERROR - Dashboard error:", str(e))
        traceback.print_exc()
        total_alumni = bsit_count = bsba_count = beed_count = 0
        employed = unemployed = self_employed = 0
        directly_related = 0
        recent_activities = []
        pending_requests = 0
        major_counts = []
    finally:
        cursor.close()
        db.close()

    return render_template(
        "dashboard.html",
        user_type=user_type,
        total_alumni=total_alumni,
        total_accounts=total_alumni,  # Using same count for simplicity
        employed=employed,
        pending_requests=pending_requests,
        recent_activities=recent_activities,
        # Program counts
        bsit_count=bsit_count,
        bsba_count=bsba_count,
        beed_count=beed_count,
        # Employment counts
        employed_count=employed,
        unemployed_count=unemployed,
        self_employed_count=self_employed,
        # Relevance count
        directly_related_count=directly_related,
        # Major counts
        major_counts=major_counts
    )




# ================= logout but =================
@app.route("/logout")
@login_required
def logout():
    log_activity("Logged out")
    session.clear()
    return redirect("/")

# ================= para sa records =================
@app.route("/records")
@login_required
def records():
    search = request.args.get("search", "").strip()
    success = request.args.get("success", "").strip()
    page = int(request.args.get("page", 1))
    per_page = 20
    offset = (page - 1) * per_page
    user_type = session.get("user_type")

    db = get_db()
    cursor = db.cursor()

    records_query, count_query, paginated_params, count_params = build_search_query(
        search, per_page=per_page, offset=offset
    )

    cursor.execute(records_query, tuple(paginated_params))
    raw_records = cursor.fetchall()

    cursor.execute(count_query, tuple(count_params))
    result = cursor.fetchone()
    total_records = result['COUNT(DISTINCT a.alumni_id)'] if result else 0

    cursor.close()
    db.close()

    total_pages = max(1, (total_records + per_page - 1) // per_page)
    start_index = offset + 1

    formatted_records = []
    for i, row in enumerate(raw_records, start=start_index):
        # Get photo path for display (using dictionary access)
        photo_path = row.get('photo', '')
        if photo_path:
            photo_url = url_for('static', filename=f'uploads/{photo_path}')
        else:
            photo_url = url_for('static', filename='images/default-avatar.svg')
        
        formatted_records.append((
            i,  # [0]  display row number
            row.get('alumni_id', ''),  # [1]  alumni_id
            row.get('stud_num', ''),  # [2]  stud_num
            photo_url,  # [3]  photo URL (2x2 display)
            row.get('last_name', ''),  # [4]  last_name
            row.get('first_name', ''),  # [5]  first_name
            row.get('middle_name', ''),  # [6]  middle_name
            row.get('address', ''),  # [7]  address
            row.get('email', ''),  # [8]  email
            row.get('contact_num', ''),  # [9]  contact_num
            row.get('added_by', ''),  # [10]  added_by
            row.get('date_added', ''),  # [11]  date_added
            row.get('program', ''),  # [12]  program
            row.get('major', ''),  # [13]  major
            row.get('employment_status', ''),  # [14]  employment_status
            row.get('job_title', ''),  # [15]  job_title
            row.get('employment_sector', ''),  # [16]  employment_sector
            row.get('degree_relevance_to_work', ''),  # [17]  work relevance
        ))

    return render_template(
        "records.html",
        records=formatted_records,
        page=page,
        total_pages=total_pages,
        search=search,
        success=success,
        user_type=user_type
    )

# ================= para sa add =================
@app.route("/add", methods=["GET", "POST"])
@login_required
def add():
    if "username" not in session:
        return redirect("/")

    # Role guard: only ALUMNI COORDINATOR can add
    if session.get("user_type") != "Alumni Coordinator":
        return "Access Denied", 403

    db = get_db()
    cursor = db.cursor()

    try:
        # Fetch programs for dropdown
        cursor.execute("SELECT program_name FROM programs ORDER BY program_name")
        programs = [row['program_name'] for row in cursor.fetchall()]
        
        if request.method == "POST":

            # ---------- Student number ----------
            stud_num_input = request.form.get("stud_num_input", "").strip()
            if not stud_num_input:
                return render_template("add.html", error="Student number is required.",
                                       user_type=session.get("user_type"), now=datetime.now())
            if not re.match(r"^\d{4}-\d{5}$", stud_num_input):
                return render_template("add.html", error="Invalid format. Example: 2023-00238",
                                       user_type=session.get("user_type"), now=datetime.now())
            stud_num = f"TAL-{stud_num_input}"

            # ---------- Names ----------
            first_name = request.form.get("first_name", "").strip()
            last_name = request.form.get("last_name", "").strip()
            middle_name = request.form.get("middle_name", "").strip()
            full_name = f"{first_name} {last_name}"

            # ---------- Contact ----------
            contact_num = re.sub(r"\D", "", request.form.get("contact_num", ""))
            if not contact_num:
                return render_template("add.html", error="Contact number is required.",
                                       user_type=session.get("user_type"), now=datetime.now())
            if len(contact_num) != 11:
                return render_template("add.html", error="Contact number must be exactly 11 digits.",
                                       user_type=session.get("user_type"), now=datetime.now())
            if not contact_num.startswith("09"):
                return render_template("add.html", error="Contact number must start with 09.",
                                       user_type=session.get("user_type"), now=datetime.now())

            # ---------- Photo ----------
            photo_file = request.files.get("photo")
            photo_filename = save_photo(photo_file)

            # ---------- INSERT ALUMNI ----------
            cursor.execute("""
                INSERT INTO alumni_table
                (stud_num, last_name, middle_name, first_name, address, email, contact_num, photo, added_by, date_added)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                stud_num,
                last_name,
                middle_name,
                first_name,
                request.form.get("address"),
                request.form.get("email"),
                contact_num,
                photo_filename,
                session.get("username"),
                date.today()
            ))
            alumni_id = cursor.lastrowid

            # ---------- DEGREE ----------
            # Validate Graduation Date
            grad_date = request.form.get("graduation_date") or None
            if grad_date:
                try:
                    grad_date_obj = datetime.strptime(grad_date, '%Y-%m-%d').date()
                    if grad_date_obj > date.today():
                        return render_template("add.html", programs=programs, error="Graduation Date cannot be greater than current date.", user_type=session.get("user_type"), now=datetime.now())
                except ValueError:
                    return render_template("add.html", programs=programs, error="Invalid Graduation Date format.", user_type=session.get("user_type"), now=datetime.now())
            
            cursor.execute("""
                INSERT INTO alumni_degree
                (alumni_id, program, major, graduation_date, added_by, date_added)
                VALUES (%s,%s,%s,%s,%s,%s)
            """, (
                alumni_id,
                request.form.get("program"),
                request.form.get("major"),
                grad_date,
                session.get("username"),
                date.today()
            ))

            # ---------- MULTI-JOB EMPLOYMENT ----------
            employment_status = request.form.get("employment_status", "")

            if employment_status == "Unemployed":
                # Single unemployed row, no jobs
                cursor.execute("""
                    INSERT INTO alumni_employment
                    (alumni_id, employment_status, employment_sector, job_title, degree_relevance_to_work, added_by, date_added)
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                """, (alumni_id, "Unemployed", None, None, None, session.get("username"), date.today()))
            else:
                job_titles = request.form.getlist("job_title[]")
                sectors = request.form.getlist("employment_sector[]")
                relevances = request.form.getlist("degree_relevance_to_work[]")

                if not job_titles:
                    # Fallback: at least one empty employed row
                    cursor.execute("""
                        INSERT INTO alumni_employment
                        (alumni_id, employment_status, employment_sector, job_title, degree_relevance_to_work, added_by, date_added)
                        VALUES (%s,%s,%s,%s,%s,%s,%s)
                    """, (alumni_id, employment_status, None, None, None, session.get("username"), date.today()))
                else:
                    for i, job_title in enumerate(job_titles):
                        sector = sectors[i] if i < len(sectors) else ""
                        relevance = relevances[i] if i < len(relevances) else ""
                        cursor.execute("""
                            INSERT INTO alumni_employment
                            (alumni_id, employment_status, employment_sector, job_title, degree_relevance_to_work, added_by, date_added)
                            VALUES (%s,%s,%s,%s,%s,%s,%s)
                        """, (
                            alumni_id,
                            employment_status,
                            sector or None,
                            job_title or None,
                            relevance or None,
                            session.get("username"),
                            date.today()
                        ))

            db.commit()
            log_activity(f"Added alumni {full_name}")
            
            cursor.close()
            db.close()
            
            # Redirect to records with success parameter
            return redirect(url_for('records', success='alumni_added'))

        return render_template("add.html", programs=programs, user_type=session.get("user_type"), now=datetime.now())

    except Exception as e:
        import traceback
        print("ERROR - Add alumni failed:", str(e))
        traceback.print_exc()
        db.rollback()
        return render_template("add.html", programs=programs, error=str(e), user_type=session.get("user_type"), now=datetime.now())

    finally:
        cursor.close()
        db.close()


# ================= pangtingin ng educ info =================
@app.route("/educ/<int:alumni_id>")
@login_required
def view_educ(alumni_id):
    search = request.args.get("search", "")
    page = request.args.get("page", 1)

    db = get_db()
    cursor = db.cursor()

    cursor.execute("""
        SELECT program, major, graduation_date
        FROM alumni_degree
        WHERE alumni_id=%s
    """, (alumni_id,))

    educ = cursor.fetchall()

    cursor.close()
    db.close()

    return render_template(
        "educ.html",
        educ=educ,
        search=request.args.get("search", ""),
        page=request.args.get("page", 1),
        user_type=session.get("user_type")
    )

# ================= pangtingin ng employment info =================
@app.route("/employ/<int:alumni_id>")
@login_required
def view_employ(alumni_id):
    search = request.args.get("search", "")
    page = request.args.get("page", 1)

    db = get_db()
    cursor = db.cursor()

    cursor.execute("""
        SELECT job_title, employment_status, employment_sector, degree_relevance_to_work
        FROM alumni_employment
        WHERE alumni_id=%s
    """, (alumni_id,))

    employ = cursor.fetchall()

    cursor.close()
    db.close()

    return render_template(
        "employ.html",
        employ=employ,
        search=request.args.get("search", ""),
        page=request.args.get("page", 1),
        user_type=session.get("user_type")
    )
@app.route("/activity")
@login_required
def activity():
    user_type = session.get("user_type")

    if user_type not in ["Admin", "Alumni Coordinator"]:
        return "Access Denied", 403

    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute("""
            SELECT username, user_type, activity, date_time
            FROM activity_logs
            ORDER BY date_time DESC
        """)
        
        logs = cursor.fetchall()
        
    except Exception as e:
        import traceback
        print("ERROR - Activity logs failed:", str(e))
        traceback.print_exc()
        logs = []
    finally:
        cursor.close()
        db.close()

    return render_template(
        "activity.html",
        logs=logs,
        user_type=user_type
    )

@app.route("/announcement", methods=["GET", "POST"])
@login_required
def announcement():
    user_type = session.get("user_type")

    if user_type not in ["Admin", "Alumni Coordinator"]:
        return "Access Denied", 403

    db = get_db()
    cursor = db.cursor()

    search = request.args.get("search", "").strip()
    selected_ids = request.args.getlist("selected_alumni")

    # ================= HANDLE POST (SEND EMAIL) =================
    success = None
    error = None

    if request.method == "POST":
        subject = request.form.get("subject")
        message_body = request.form.get("message")
        selected_ids = request.form.getlist("selected_alumni")

        try:
            # For Admin: Handle approval actions only (no alumni selection required)
            if user_type == "Admin":
                approval_action = request.form.get("approval_action")
                request_id = request.form.get("request_id")
                
                if approval_action and request_id:
                    if approval_action == "approve":
                        new_status = "approved"
                    elif approval_action == "reject":
                        new_status = "rejected"
                    else:
                        error = "Invalid approval action"
                    
                    if not error:
                        cursor.execute("""
                            UPDATE announcement_requests 
                            SET status = %s, admin_note = %s, approved_at = %s, approved_by = %s
                            WHERE id = %s
                        """, (
                            new_status,
                            request.form.get("admin_note", ""),
                            date.today(),
                            session.get("username"),
                            request_id
                        ))
                        
                        db.commit()
                        
                        # Send the actual announcement if approved
                        if approval_action == "approve":
                            # Get the original request details
                            cursor.execute("""
                                SELECT subject, message, recipient_emails
                                FROM announcement_requests
                                WHERE id = %s
                            """, (request_id,))
                            
                            req = cursor.fetchone()
                            
                            if req and req['recipient_emails']:
                                recipients = req['recipient_emails'].split(',')
                                msg = Message(
                                    subject=req['subject'],
                                    recipients=recipients
                                )
                                msg.body = req['message']
                                mail.send(msg)
                                success = f"Announcement approved and sent to {len(recipients)} alumni."
                            else:
                                error = "No valid email recipients found for this announcement."
                        else:
                            success = f"Announcement request rejected successfully."
                else:
                    error = "Missing approval action or request ID"
            
            # For Alumni Coordinator: Handle new announcement submission
            elif user_type == "Alumni Coordinator":
                if not selected_ids:
                    error = "Please select at least one alumni."
                elif not subject or not message_body:
                    error = "Subject and message are required."
                else:
                    placeholders = ','.join(['%s'] * len(selected_ids))
                    cursor.execute(f"""
                        SELECT email FROM alumni_table 
                        WHERE alumni_id IN ({placeholders})
                        AND email IS NOT NULL
                        AND TRIM(email) != ''
                    """, selected_ids)

                    recipients = [r["email"] for r in cursor.fetchall()]

                    if recipients:
                        # Create approval request instead of sending immediately
                        cursor.execute("""
                            INSERT INTO announcement_requests 
                            (coordinator_id, subject, message, recipient_emails, status, created_at)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (
                            session.get("username"),
                            subject,
                            message_body,
                            ",".join(recipients),
                            "pending",
                            date.today()
                        ))
                        
                        db.commit()
                        success = f"Announcement request submitted for approval to {len(recipients)} alumni."
                    else:
                        error = "No valid email recipients found."

        except Exception as e:
            error = str(e)

    # ================= GET REQUESTS (FOR Admin) =================
    if user_type == "Admin":
        # Get pending approval requests
        cursor.execute("""
            SELECT id, coordinator_id, subject, message, recipient_emails, status, created_at
            FROM announcement_requests
            WHERE status = 'pending'
            ORDER BY created_at DESC
        """)
        pending_requests = cursor.fetchall()
        
        try:
            # Get approved history
            cursor.execute("""
                SELECT id, coordinator_id, subject, message, recipient_emails, status, created_at, approved_at, approved_by, admin_note
                FROM announcement_requests
                WHERE status = 'approved'
                ORDER BY approved_at DESC
                LIMIT 20
            """)
            approved_history = cursor.fetchall()
            
            # Get rejected history
            cursor.execute("""
                SELECT id, coordinator_id, subject, message, recipient_emails, status, created_at, approved_at, approved_by, admin_note
                FROM announcement_requests
                WHERE status = 'rejected'
                ORDER BY approved_at DESC
                LIMIT 20
            """)
            rejected_history = cursor.fetchall()
            
        except Exception as e:
            import traceback
            print("ERROR - Announcement failed:", str(e))
            traceback.print_exc()
            # Set defaults on error
            pending_requests = []
            approved_history = []
            rejected_history = []
            alumni_list = []
            my_requests = []
        finally:
            cursor.close()
            db.close()

        return render_template(
            "announcement.html",
            pending_requests=pending_requests,
            approved_history=approved_history,
            rejected_history=rejected_history,
            success=success,
            error=error,
            search=search,
            selected_ids=selected_ids,
            user_type=user_type
        )
    
    # ================= ORIGINAL FUNCTIONALITY (FOR COORDINATOR) =================
    # Get advanced filter params
    filter_program           = request.args.get("filter_program", "").strip()
    filter_major             = request.args.get("filter_major", "").strip()
    filter_lastname          = request.args.get("filter_lastname", "").strip()
    filter_firstname         = request.args.get("filter_firstname", "").strip()
    filter_address           = request.args.get("filter_address", "").strip()
    filter_contact           = request.args.get("filter_contact", "").strip()
    filter_employment_status = request.args.get("filter_employment_status", "").strip()
    filter_sector            = request.args.get("filter_sector", "").strip()
    filter_job_title         = request.args.get("filter_job_title", "").strip()

    filter_contact_clean = re.sub(r"\D", "", filter_contact)

    # ================= BUILD ALUMNI QUERY WITH FILTERS =================
    query = """
        SELECT
            a.alumni_id,
            a.stud_num,
            CONCAT(a.first_name, ' ', a.last_name) AS full_name,
            a.last_name,
            a.first_name,
            a.middle_name,
            a.address,
            a.email,
            a.contact_num,
            d.program,
            d.major,
            d.graduation_date,
            e.employment_status,
            e.employment_sector,
            e.job_title,
            e.degree_relevance_to_work
        FROM alumni_table a
        LEFT JOIN alumni_degree d ON a.alumni_id = d.alumni_id
        LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
        WHERE a.email IS NOT NULL
          AND TRIM(a.email) != ''
    """

    params = []

    # Enhanced search - search ALL alumni fields
    if search:
        like = f"%{search.lower()}%"
        query += """
            AND (
                LOWER(COALESCE(a.stud_num, '')) LIKE %s
                OR LOWER(COALESCE(a.first_name, '')) LIKE %s
                OR LOWER(COALESCE(a.last_name, '')) LIKE %s
                OR LOWER(COALESCE(a.middle_name, '')) LIKE %s
                OR LOWER(COALESCE(a.address, '')) LIKE %s
                OR LOWER(COALESCE(a.email, '')) LIKE %s
                OR LOWER(COALESCE(a.contact_num, '')) LIKE %s
                OR LOWER(COALESCE(d.program, '')) LIKE %s
                OR LOWER(COALESCE(d.major, '')) LIKE %s
                OR LOWER(COALESCE(e.employment_status, '')) LIKE %s
                OR LOWER(COALESCE(e.employment_sector, '')) LIKE %s
                OR LOWER(COALESCE(e.job_title, '')) LIKE %s
                OR LOWER(COALESCE(e.degree_relevance_to_work, '')) LIKE %s
            )
        """
        params += [like, like, like, like, like, like, like, like, like, like, like, like, like]

    # Advanced filters
    if filter_program:
        query += " AND LOWER(COALESCE(d.program, '')) LIKE %s"
        params.append(f"%{filter_program.lower()}%")

    if filter_major:
        query += " AND LOWER(COALESCE(d.major, '')) LIKE %s"
        params.append(f"%{filter_major.lower()}%")

    if filter_lastname:
        query += " AND LOWER(COALESCE(a.last_name, '')) LIKE %s"
        params.append(f"%{filter_lastname.lower()}%")

    if filter_firstname:
        query += " AND LOWER(COALESCE(a.first_name, '')) LIKE %s"
        params.append(f"%{filter_firstname.lower()}%")

    if filter_address:
        query += " AND LOWER(COALESCE(a.address, '')) LIKE %s"
        params.append(f"%{filter_address.lower()}%")

    if filter_contact_clean:
        query += """
            AND REPLACE(REPLACE(COALESCE(CAST(a.contact_num AS CHAR), ''), ' ', ''), '-', '') LIKE %s
        """
        params.append(f"%{filter_contact_clean}%")

    if filter_employment_status:
        query += " AND LOWER(COALESCE(e.employment_status, '')) = %s"
        params.append(filter_employment_status.lower())

    if filter_sector:
        query += " AND LOWER(COALESCE(e.employment_sector, '')) LIKE %s"
        params.append(f"%{filter_sector.lower()}%")

    if filter_job_title:
        query += " AND LOWER(COALESCE(e.job_title, '')) LIKE %s"
        params.append(f"%{filter_job_title.lower()}%")

    query += " GROUP BY a.alumni_id ORDER BY a.first_name ASC"

    cursor.execute(query, tuple(params))
    alumni_list = cursor.fetchall()
    # Get coordinator's own announcement requests
    cursor.execute("""
        SELECT id, subject, message, recipient_emails, status, created_at, approved_at, approved_by, admin_note
        FROM announcement_requests
        WHERE coordinator_id = %s
        ORDER BY created_at DESC
        LIMIT 10
    """, (session.get("username"),))
    my_requests = cursor.fetchall()

    cursor.close()
    db.close()

    return render_template(
        "announcement.html",
        alumni_list=alumni_list,
        my_requests=my_requests,
        success=success,
        error=error,
        search=search,
        selected_ids=selected_ids,
        user_type=user_type
    )



@app.route("/generate-record/<int:alumni_id>/<file_type>")
@coordinator_required
def generate_single_record(alumni_id, file_type):
    if "username" not in session:
        return redirect("/")

    db = get_db()
    cursor = db.cursor()

    cursor.execute("""
        SELECT 
            a.alumni_id, a.stud_num, a.last_name, a.first_name, a.middle_name,
            a.address, a.email, a.contact_num, a.added_by, a.date_added,
            d.program, d.major, d.graduation_date,
            e.employment_status, e.job_title, e.employment_sector, e.degree_relevance_to_work
        FROM alumni_table a
        LEFT JOIN alumni_degree d ON a.alumni_id = d.alumni_id
        LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
        WHERE a.alumni_id = %s
    """, (alumni_id,))

    r = cursor.fetchone()
    cursor.close()
    db.close()

    if not r:
        return "Record not found"

    def safe(v):
        return v if v is not None else ""

    # ================= ONLY FIX: ENSURE DATA ORDER IS CORRECT =================
    fixed_r = list(r)

    # normalize employment fields (IMPORTANT FIX)
    employment_status = safe(fixed_r[14])
    job_title = safe(fixed_r[15])

    # ensure valid fallback (optional safety)
    if employment_status == "":
        employment_status = "Unemployed"

    fixed_r[14] = employment_status   # Status
    fixed_r[15] = job_title           # Job Title

    fixed_r = tuple(fixed_r)

    # ================= PDF (NO DESIGN CHANGE) =================
    if file_type == "pdf":
        buffer = BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=LETTER,
            rightMargin=70,
            leftMargin=70,
            topMargin=65,
            bottomMargin=65
        )

        elements = []

        # 🔥 DO NOT CHANGE THIS (PRESERVES YOUR DESIGN)
        build_alumni_page(elements, fixed_r)

        doc.build(elements, onFirstPage=draw_layout)

        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"alumni_{alumni_id}.pdf",
            mimetype="application/pdf"
        )

    # ================= EXCEL (UNCHANGED STRUCTURE) =================
    elif file_type == "excel":
        wb = Workbook()
        ws = wb.active

        ws.append([
            "ID","Student No","Last Name","First Name","Middle Name",
            "Address","Email","Contact",
            "Program","Major","Admission","Graduation",
            "Status", "Job Title","Sector","Relevance"
        ])

        ws.append([
            safe(r[0]), safe(r[1]), safe(r[2]), safe(r[3]), safe(r[4]),
            safe(r[5]), safe(r[6]), safe(r[7]),
            safe(r[10]), safe(r[11]),
            safe(r[12]), safe(r[13]),
            safe(fixed_r[14]), safe(fixed_r[15]), safe(r[16]), safe(r[17])
        ])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"alumni_{alumni_id}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    return "Invalid file type"


# ================= Para sa update =================
@app.route("/update/<int:alumni_id>", methods=["GET", "POST"])
@coordinator_required
def update(alumni_id):
    if "username" not in session:
        return redirect("/")

    db = get_db()
    cursor = db.cursor()

    try:
        # Fetch programs for dropdown
        cursor.execute("SELECT program_name FROM programs ORDER BY program_name")
        programs = [row['program_name'] for row in cursor.fetchall()]
        
        cursor.execute("SELECT * FROM alumni_table WHERE alumni_id=%s", (alumni_id,))
        alumni = cursor.fetchone()

        cursor.execute("SELECT * FROM alumni_degree WHERE alumni_id=%s LIMIT 1", (alumni_id,))
        educ = cursor.fetchone()

        # Fetch ALL employment rows for this alumni
        cursor.execute("SELECT * FROM alumni_employment WHERE alumni_id=%s ORDER BY alumni_id ASC", (alumni_id,))
        employ = cursor.fetchall()

        if request.method == "POST":

            # ---------- Student number ----------
            stud_num = request.form.get("stud_num", "").strip()
            if stud_num:
                stud_num = stud_num.replace("TAL-", "")
                stud_num = f"TAL-{stud_num.replace(' ', '')}"
            else:
                stud_num = None

            # ---------- Contact ----------
            contact = re.sub(r"\D", "", request.form.get("contact_num", ""))
            if not contact:
                return render_template("update.html", programs=programs, alumni=alumni, educ=educ, employ=employ,
                                       error="Contact number is required.", user_type=session.get("user_type"), now=datetime.now())
            if len(contact) != 11:
                return render_template("update.html", programs=programs, alumni=alumni, educ=educ, employ=employ,
                                       error="Contact number must be exactly 11 digits.",
                                       user_type=session.get("user_type"), now=datetime.now())
            if not contact.startswith("09"):
                return render_template("update.html", programs=programs, alumni=alumni, educ=educ, employ=employ,
                                       error="Contact number must start with 09.", user_type=session.get("user_type"), now=datetime.now())

            # ---------- Photo ----------
            old_photo = alumni.get("photo") if alumni else None
            remove_flag = request.form.get("remove_photo", "0")
            photo_file = request.files.get("photo")
            new_photo = old_photo  # default: keep existing

            if remove_flag == "1":
                delete_photo(old_photo)
                new_photo = None
            elif photo_file and photo_file.filename:
                saved = save_photo(photo_file)
                if saved:
                    delete_photo(old_photo)  # remove old one
                    new_photo = saved

            # ---------- UPDATE PERSONAL ----------
            cursor.execute("""
                UPDATE alumni_table
                SET stud_num=%s, last_name=%s, middle_name=%s, first_name=%s,
                    address=%s, email=%s, contact_num=%s, photo=%s
                WHERE alumni_id=%s
            """, (
                stud_num,
                request.form.get("last_name"),
                request.form.get("middle_name"),
                request.form.get("first_name"),
                request.form.get("address"),
                request.form.get("email"),
                contact,
                new_photo,
                alumni_id
            ))

            # ---------- UPDATE DEGREE ----------
            # Validate Graduation Date
            grad_date = request.form.get("graduation_date") or None
            if grad_date:
                try:
                    grad_date_obj = datetime.strptime(grad_date, '%Y-%m-%d').date()
                    if grad_date_obj > date.today():
                        return render_template("update.html", programs=programs, alumni=alumni, educ=educ, employ=employ,
                                               error="Graduation Date cannot be greater than current date.", user_type=session.get("user_type"), now=datetime.now())
                except ValueError:
                    return render_template("update.html", programs=programs, alumni=alumni, educ=educ, employ=employ,
                                               error="Invalid Graduation Date format.", user_type=session.get("user_type"), now=datetime.now())
            
            cursor.execute("""
                UPDATE alumni_degree
                SET program=%s, major=%s, graduation_date=%s
                WHERE alumni_id=%s
            """, (
                request.form.get("program"),
                request.form.get("major"),
                grad_date,
                alumni_id
            ))

            # ---------- REPLACE ALL EMPLOYMENT ROWS ----------
            cursor.execute("DELETE FROM alumni_employment WHERE alumni_id=%s", (alumni_id,))

            employment_status = request.form.get("employment_status", "")

            if employment_status == "Unemployed":
                cursor.execute("""
                    INSERT INTO alumni_employment
                    (alumni_id, employment_status, employment_sector, job_title, degree_relevance_to_work, added_by, date_added)
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                """, (alumni_id, "Unemployed", None, None, None, session.get("username"), date.today()))
            else:
                job_titles = request.form.getlist("job_title[]")
                sectors = request.form.getlist("employment_sector[]")
                relevances = request.form.getlist("degree_relevance_to_work[]")

                if not job_titles:
                    cursor.execute("""
                        INSERT INTO alumni_employment
                        (alumni_id, employment_status, employment_sector, job_title, degree_relevance_to_work, added_by, date_added)
                        VALUES (%s,%s,%s,%s,%s,%s,%s)
                    """, (alumni_id, employment_status, None, None, None, session.get("username"), date.today()))
                else:
                    for i, job_title in enumerate(job_titles):
                        sector = sectors[i] if i < len(sectors) else ""
                        relevance = relevances[i] if i < len(relevances) else ""
                        cursor.execute("""
                            INSERT INTO alumni_employment
                            (alumni_id, employment_status, employment_sector, job_title, degree_relevance_to_work, added_by, date_added)
                            VALUES (%s,%s,%s,%s,%s,%s,%s)
                        """, (
                            alumni_id,
                            employment_status,
                            sector or None,
                            job_title or None,
                            relevance or None,
                            session.get("username"),
                            date.today()
                        ))

            db.commit()

            # Log
            cursor.execute("SELECT stud_num FROM alumni_table WHERE alumni_id=%s", (alumni_id,))
            stud = cursor.fetchone()
            log_activity(f"Updated alumni {stud['stud_num'] if stud else alumni_id}")

            flash("Alumni Successfully updated", "success")
            return redirect("/records")

        # ---------- GET: strip TAL- prefix for form ----------
        if alumni and alumni.get("stud_num"):
            alumni["stud_num"] = alumni["stud_num"].replace("TAL-", "")

        return render_template(
            "update.html",
            programs=programs,
            alumni=alumni,
            educ=educ,
            employ=employ,
            user_type=session.get("user_type"),
            now=datetime.now()
        )

    except Exception as e:
        import traceback
        print("ERROR - Update alumni failed:", str(e))
        traceback.print_exc()
        db.rollback()
        return str(e)

    finally:
        cursor.close()
        db.close()


# ================= Para sa ARCHIVE =================
@app.route("/archive/<int:alumni_id>", methods=["POST"])
@coordinator_required
def archive(alumni_id):

    db = get_db()
    cursor = db.cursor()

    try:

        cursor.execute("""
            SELECT a.alumni_id, a.stud_num, a.last_name, a.first_name, a.middle_name,
                   a.address, a.email,
                   d.program, d.major, d.graduation_date,
                   e.employment_status, e.employment_sector, e.job_title, e.degree_relevance_to_work
            FROM alumni_table a
            LEFT JOIN alumni_degree d ON a.alumni_id = d.alumni_id
            LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
            WHERE a.alumni_id = %s
        """, (alumni_id,))

        r = cursor.fetchone()

        if not r or r.get("alumni_id") is None:
            db.rollback()
            return "Record not found", 404

        # ================= INSERT ARCHIVE =================
        cursor.execute("""
            INSERT INTO archive (
                alumni_id, stud_num, last_name, first_name, middle_name,
                address, email,
                program, major, graduation_date,
                employment_status, employment_sector, job_title, degree_relevance_to_work,
                archived_by, date_archived
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            safe(r.get("alumni_id")),
            safe(r.get("stud_num")),
            safe(r.get("last_name")),
            safe(r.get("first_name")),
            safe(r.get("middle_name")),
            safe(r.get("address")),
            safe(r.get("email")),
            safe(r.get("program")),
            safe(r.get("major")),
            safe_date(r.get("graduation_date")),
            safe(r.get("employment_status")),
            safe(r.get("employment_sector")),
            safe(r.get("job_title")),
            safe(r.get("degree_relevance_to_work")),
            session.get("username"),
            date.today()
        ))

        cursor.execute("DELETE FROM alumni_degree WHERE alumni_id = %s", (alumni_id,))
        cursor.execute("DELETE FROM alumni_employment WHERE alumni_id = %s", (alumni_id,))
        cursor.execute("DELETE FROM alumni_table WHERE alumni_id = %s", (alumni_id,))

        db.commit()

        try:
            resequence_alumni_ids()
        except Exception as e:
            print("Resequence warning:", e)

        full_name = f"{r.get('last_name')}, {r.get('first_name')}"
        if r.get("middle_name"):
            full_name += f" {r.get('middle_name')}"
        log_activity(f"Archived alumni: {full_name}")

        # Check if this is an AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({"success": True, "message": "Alumni record archived successfully."})

        # For regular form submission, redirect back to records
        return redirect(url_for('records'))

    except Exception as e:
        import traceback
        print("ERROR - Archive failed:", str(e))
        traceback.print_exc()
        db.rollback()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({"success": False, "error": str(e)})
        return f"Archive error: {str(e)}"

    finally:
        cursor.close()
        db.close()

    return redirect(request.referrer or "/records")

# ================= BULK IMPORT FUNCTIONALITY =================
# Configure upload settings
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def normalize_column_name(col_name):
    """Normalize column names to match database fields"""
    col_name = str(col_name).strip().lower()
    
    # Personal Info mappings - more specific matching
    if any(keyword in col_name for keyword in ['student number', 'stud_num', 'student no']):
        return 'stud_num'
    elif any(keyword in col_name for keyword in ['last name', 'surname', 'family', 'last']):
        return 'last_name'
    elif any(keyword in col_name for keyword in ['first name', 'given', 'first']):
        return 'first_name'
    elif any(keyword in col_name for keyword in ['middle name', 'middle']):
        return 'middle_name'
    elif any(keyword in col_name for keyword in ['email address', 'email']):
        return 'email'
    elif any(keyword in col_name for keyword in ['address']):
        return 'address'
    elif any(keyword in col_name for keyword in ['contact number', 'phone', 'mobile', 'contact']):
        return 'contact_num'
    elif any(keyword in col_name for keyword in ['photo']):
        return 'photo'
    
    # Educational Info mappings
    elif any(keyword in col_name for keyword in ['program', 'course', 'degree']):
        return 'program'
    elif any(keyword in col_name for keyword in ['major', 'specialization']):
        return 'major'
    elif any(keyword in col_name for keyword in ['graduation', 'grad']):
        return 'graduation_date'
    
    # Employment Info mappings
    elif any(keyword in col_name for keyword in ['employment', 'status']):
        return 'employment_status'
    elif any(keyword in col_name for keyword in ['sector']):
        return 'employment_sector'
    elif any(keyword in col_name for keyword in ['job', 'position', 'title']):
        return 'job_title'
    elif any(keyword in col_name for keyword in ['relevance', 'relevant']):
        return 'degree_relevance_to_work'
    
    return col_name

def validate_student_number(stud_num):
    """Validate and format student number"""
    if not stud_num:
        return None, "Student number is required"
    
    stud_num = str(stud_num).strip()
    
    # Remove TAL- prefix if present
    if stud_num.startswith('TAL-'):
        stud_num = stud_num[4:]
    
    # Check format: YYYY-NNNNN
    if not re.match(r'^\d{4}-\d{5}$', stud_num):
        return None, "Invalid student number format. Expected: YYYY-NNNNN (e.g., 2023-00238)"
    
    return f"TAL-{stud_num}", None

def validate_email(email):
    """Validate email format"""
    if not email:
        return None, "Email is required"
    
    email = str(email).strip()
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    
    if not re.match(email_pattern, email):
        return None, "Invalid email format"
    
    return email, None

def validate_date(date_str):
    """Validate and parse date"""
    if not date_str:
        return None, None
    
    try:
        if isinstance(date_str, datetime):
            return date_str.date(), None
        
        # Try different date formats
        for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']:
            try:
                return datetime.strptime(str(date_str).strip(), fmt).date(), None
            except ValueError:
                continue
        
        return None, "Invalid date format. Use YYYY-MM-DD, MM/DD/YYYY, or DD/MM/YYYY"
    except Exception:
        return None, "Invalid date format"

def parse_file(file_path):
    """Parse Excel or CSV file and return data"""
    try:
        file_ext = file_path.rsplit('.', 1)[1].lower()
        
        if file_ext in ['xlsx', 'xls']:
            # Parse Excel file
            df = pd.read_excel(file_path, engine='openpyxl')
        elif file_ext == 'csv':
            # Parse CSV file
            df = pd.read_csv(file_path)
        else:
            return None, "Unsupported file format"
        
        # Remove empty rows
        df = df.dropna(how='all')
        
        return df, None
    
    except Exception as e:
        return None, f"Error parsing file: {str(e)}"

def validate_import_data(df):
    """Validate and map import data"""
    errors = []
    warnings = []
    valid_records = []
    
    # Normalize column names
    df.columns = [normalize_column_name(col) for col in df.columns]
    
    required_fields = ['stud_num', 'last_name', 'first_name', 'email']
    missing_fields = [field for field in required_fields if field not in df.columns]
    
    if missing_fields:
        errors.append(f"Missing required columns: {', '.join(missing_fields)}")
        return valid_records, errors, warnings
    
    # Check for duplicate student numbers
    if 'stud_num' in df.columns:
        duplicate_stud_nums = df['stud_num'].duplicated()
        if duplicate_stud_nums.any():
            warnings.append(f"Found duplicate student numbers in file")
    
    # Process each row
    for index, row in df.iterrows():
        record = {}
        row_errors = []
        
        # Validate and process each field
        for field in required_fields:
            value = row.get(field, '')
            if pd.isna(value) or str(value).strip() == '':
                row_errors.append(f"{field} is required")
                continue
            
            if field == 'stud_num':
                formatted_stud_num, error = validate_student_number(value)
                if error:
                    row_errors.append(error)
                else:
                    record[field] = formatted_stud_num
            elif field == 'email':
                formatted_email, error = validate_email(value)
                if error:
                    row_errors.append(error)
                else:
                    record[field] = formatted_email
            else:
                record[field] = str(value).strip()
        
        # Process optional fields
        optional_fields = ['middle_name', 'address', 'contact_num', 'photo', 
                          'program', 'major', 'graduation_date',
                          'employment_status', 'employment_sector', 'job_title', 
                          'degree_relevance_to_work']
        
        for field in optional_fields:
            if field in df.columns:
                value = row.get(field, '')
                if pd.notna(value) and str(value).strip() != '':
                    if field == 'graduation_date':
                        formatted_date, error = validate_date(value)
                        if error:
                            row_errors.append(f"{field}: {error}")
                        else:
                            record[field] = formatted_date
                    else:
                        record[field] = str(value).strip()
        
        if row_errors:
            errors.append(f"Row {index + 2}: {'; '.join(row_errors)}")
        else:
            valid_records.append(record)
    
    return valid_records, errors, warnings

@app.route("/import-alumni", methods=["GET", "POST"])
@login_required
def import_alumni():
    # Role guard: only Alumni Coordinator can import
    if session.get("user_type") != "Alumni Coordinator":
        return "Access Denied", 403
    
    if request.method == "GET":
        return render_template("import_alumni.html", user_type=session.get("user_type"))
    
    # Handle file upload
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "No file selected"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "error": "No file selected"}), 400
    
    if not allowed_file(file.filename):
        return jsonify({"success": False, "error": "Invalid file type. Please upload Excel (.xlsx, .xls) or CSV (.csv) files"}), 400
    
    try:
        # Create upload directory if it doesn't exist
        if not os.path.exists(app.config['UPLOAD_FOLDER']):
            os.makedirs(app.config['UPLOAD_FOLDER'])
        
        # Save uploaded file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"import_{timestamp}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Parse and validate file
        df, parse_error = parse_file(file_path)
        if parse_error:
            return jsonify({"success": False, "error": parse_error}), 400
        
        # Validate data
        valid_records, errors, warnings = validate_import_data(df)
        
        if errors:
            return jsonify({
                "success": False, 
                "error": "Validation errors found",
                "validation_errors": errors,
                "warnings": warnings
            }), 400
        
        if not valid_records:
            return jsonify({"success": False, "error": "No valid records found in file"}), 400
        
        # Return preview data
        preview_data = []
        for i, record in enumerate(valid_records[:10]):  # Show first 10 records
            preview_data.append({
                "row": i + 1,
                "stud_num": record.get('stud_num', ''),
                "last_name": record.get('last_name', ''),
                "first_name": record.get('first_name', ''),
                "email": record.get('email', ''),
                "program": record.get('program', ''),
                "major": record.get('major', ''),
                "employment_status": record.get('employment_status', '')
            })
        
        return jsonify({
            "success": True,
            "message": f"File processed successfully. Found {len(valid_records)} valid records.",
            "total_records": len(valid_records),
            "preview": preview_data,
            "warnings": warnings,
            "file_path": file_path
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": f"Error processing file: {str(e)}"}), 500

@app.route("/confirm-import", methods=["POST"])
@login_required
def confirm_import():
    # Role guard: only Alumni Coordinator can import
    if session.get("user_type") != "Alumni Coordinator":
        return "Access Denied", 403
    
    try:
        data = request.get_json()
        file_path = data.get('file_path')
        
        if not file_path or not os.path.exists(file_path):
            return jsonify({"success": False, "error": "File not found"}), 400
        
        # Parse and validate file again
        df, parse_error = parse_file(file_path)
        if parse_error:
            return jsonify({"success": False, "error": parse_error}), 400
        
        valid_records, errors, warnings = validate_import_data(df)
        
        if errors:
            return jsonify({
                "success": False, 
                "error": "Validation errors found",
                "validation_errors": errors
            }), 400
        
        # Import records to database
        db = get_db()
        cursor = db.cursor()
        
        imported_count = 0
        skipped_count = 0
        import_errors = []
        
        for i, record in enumerate(valid_records):
            try:
                # Check for duplicate student number
                cursor.execute("SELECT alumni_id FROM alumni_table WHERE stud_num = %s", 
                             (record['stud_num'],))
                existing = cursor.fetchone()
                
                if existing:
                    skipped_count += 1
                    continue
                
                # Start transaction for this record
                # PyMySQL uses autocommit=False, no explicit start_transaction needed
                
                # Insert into alumni_table
                cursor.execute("""
                    INSERT INTO alumni_table
                    (stud_num, last_name, middle_name, first_name, address, email, contact_num, photo, added_by, date_added)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    record.get('stud_num', ''),
                    record.get('last_name', ''),
                    record.get('middle_name', ''),
                    record.get('first_name', ''),
                    record.get('address', ''),
                    record.get('email', ''),
                    record.get('contact_num', ''),
                    record.get('photo', ''),
                    session.get("username"),
                    date.today()
                ))
                
                alumni_id = cursor.lastrowid
                
                # Insert into alumni_degree if educational data exists
                if any(record.get(field) for field in ['program', 'major', 'graduation_date']):
                    cursor.execute("""
                        INSERT INTO alumni_degree
                        (alumni_id, program, major, graduation_date, added_by, date_added)
                        VALUES (%s,%s,%s,%s,%s,%s)
                    """, (
                        alumni_id,
                        record.get('program', ''),
                        record.get('major', ''),
                        record.get('graduation_date'),
                        session.get("username"),
                        date.today()
                    ))
                
                # Insert into alumni_employment if employment data exists
                if any(record.get(field) for field in ['employment_status', 'employment_sector', 'job_title', 'degree_relevance_to_work']):
                    cursor.execute("""
                        INSERT INTO alumni_employment
                        (alumni_id, employment_status, employment_sector, job_title, degree_relevance_to_work, added_by, date_added)
                        VALUES (%s,%s,%s,%s,%s,%s,%s)
                    """, (
                        alumni_id,
                        record.get('employment_status', ''),
                        record.get('employment_sector', ''),
                        record.get('job_title', ''),
                        record.get('degree_relevance_to_work', ''),
                        session.get("username"),
                        date.today()
                    ))
                
                db.commit()
                imported_count += 1
                
            except Exception as e:
                db.rollback()
                import_errors.append(f"Record {i+1}: {str(e)}")
        
        cursor.close()
        db.close()
        
        # Clean up uploaded file
        try:
            os.remove(file_path)
        except:
            pass
        
        return jsonify({
            "success": True,
            "message": f"Import completed successfully. Imported {imported_count} records.",
            "imported_count": imported_count,
            "skipped_count": skipped_count,
            "import_errors": import_errors,
            "warnings": warnings
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": f"Import failed: {str(e)}"}), 500

@app.route("/download-import-template")
@login_required
def download_import_template():
    # Role guard: only Alumni Coordinator can download template
    if session.get("user_type") != "Alumni Coordinator":
        return "Access Denied", 403
    
    try:
        # Create Excel template
        wb = Workbook()
        ws = wb.active
        ws.title = "Alumni Import Template"
        
        # Add headers
        headers = [
            'Student Number', 'Last Name', 'First Name', 'Middle Name', 'Address', 'Email', 'Contact Number',
            'Program', 'Major', 'Date of Admission', 'Graduation Date',
            'Employment Status', 'Employment Sector', 'Job Title', 'Degree Relevance to Work'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # Add sample data
        sample_data = [
            '2023-00238', 'Dela Cruz', 'Juan', 'Santos', '123 Main St, Bantug, Nueva Ecija', 
            'juan.delacruz@email.com', '09171234567', 'BSIT', 'Web Development', '2019-06-01', '2023-05-15',
            'Employed', 'Private', 'Software Developer', 'Very Relevant'
        ]
        
        for col, value in enumerate(sample_data, 1):
            ws.cell(row=2, column=col, value=value)
        
        # Add instructions
        instructions = [
            "INSTRUCTIONS:",
            "1. Required fields: Student Number, Last Name, First Name, Email",
            "2. Student Number format: YYYY-NNNNN (e.g., 2023-00238)",
            "3. Date format: YYYY-MM-DD (e.g., 2023-05-15)",
            "4. Remove sample data before importing",
            "5. Do not modify header row"
        ]
        
        for i, instruction in enumerate(instructions, 4):
            ws.cell(row=i, column=1, value=instruction)
            ws.cell(row=i, column=1).font = Font(bold=True, color="FF0000")
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name='alumni_import_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return f"Error generating template: {str(e)}", 500



# ================= Para sa graduates =================
@app.route("/api/combined_graduates")
@login_required
def combined_graduates():
    if "username" not in session:
        return {"labels": [], "datasets": []}

    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("""
            SELECT 
                YEAR(d.graduation_date) AS year,
                d.program,
                d.major,
                COUNT(DISTINCT a.alumni_id) AS total
            FROM alumni_table a
            INNER JOIN alumni_degree d ON a.alumni_id = d.alumni_id
            WHERE d.graduation_date IS NOT NULL
              AND d.program IS NOT NULL
            GROUP BY year, d.program, d.major
            HAVING total > 0
            ORDER BY year ASC, d.program ASC
        """)
        rows = cursor.fetchall()
        
        # Process data for Chart.js
        years = []
        datasets = []
        
        # Get unique years and programs
        all_years = sorted(set(row['year'] for row in rows if row['year'] is not None))
        all_programs = sorted(set(row['program'] for row in rows if row['program'] is not None))
        
        # Create datasets for each program
        for program in all_programs:
            program_data = []
            for year in all_years:
                # Find the total for this program and year
                total = 0
                for row in rows:
                    if row['year'] == year and row['program'] == program:
                        total = row['total']
                        break
                program_data.append(total)
            
            datasets.append({
                'label': program,
                'data': program_data
            })
        
        cursor.close()
        db.close()
        
        return {"labels": all_years, "datasets": datasets}
        
    except Exception as e:
        print("Combined graduates error:", e)
        cursor.close()
        db.close()
        return {"labels": [], "datasets": []}


#overall graduates
@app.route("/api/program_summary")
@login_required
def program_summary():
    db = get_db()
    cursor = db.cursor()

    cursor.execute("""
        SELECT 
            d.program,
            d.major,
            COUNT(DISTINCT a.alumni_id) AS total
        FROM alumni_table a
        INNER JOIN alumni_degree d ON a.alumni_id = d.alumni_id
        WHERE d.program IS NOT NULL
        GROUP BY d.program, d.major
        ORDER BY d.program, d.major
    """)

    rows = cursor.fetchall()
    cursor.close()
    db.close()

    data = {}

    for r in rows:
        program = r["program"]
        major = r["major"]
        total = r["total"]

        if program not in data:
            data[program] = {
                "total": 0,
                "majors": {}
            }

        data[program]["total"] += total
        data[program]["majors"][major] = total

    return data
#====Creation ng User Account=====#

@app.route("/register-admin.html")
def register_admin():
    """Standalone Admin registration page"""
    return render_template("register_admin.html", error=None, success=None)

@app.route("/register-coordinator.html")
def register_coordinator():
    """Standalone Alumni Coordinator registration page"""
    return render_template("register_coordinator.html", error=None, success=None)

@app.route("/createuser.html")
def createuser_public():
    """Public route for Admin/Coordinator registration"""
    return render_template("createuser.html", user_type=None, users=None, error=None, success=None)

@app.route("/create-user")
@admin_required
def create_user():

    db = get_db()
    cursor = db.cursor()

    # GET ALL USERS
    cursor.execute("""
        SELECT id, Lname, Fname, Mname
        FROM user_information
        ORDER BY id DESC
    """)
    users = cursor.fetchall()

    selected_user = None

    if users:
        first_id = users[0]['id'] if users else None

        # USER INFO
        cursor.execute("""
            SELECT address, email, contact, addedBy, dateAdded
            FROM user_information
            WHERE id = %s
        """, (first_id,))
        info = cursor.fetchone()

        # LOGIN INFO (FIXED MATCH ID)
        cursor.execute("""
            SELECT username, password, user_type
            FROM user_login
            WHERE id = %s
        """, (first_id,))
        login = cursor.fetchone()

        selected_user = {
            "info": info or {},
            "login": login or {}
        }

    cursor.close()
    db.close()

    return render_template(
        "create.html",
        users=users,
        selected_user=selected_user,
        user_type=session.get("user_type")
    )


@app.route("/create-user-form")
@admin_required
def create_user_form():
    return render_template(
        "createuser.html",
        user_type=session.get("user_type")
    )


@app.route("/save-user-public", methods=["POST"])
def save_user_public():
    """Public registration route for Admin/Coordinator accounts"""
    error = None
    success = None
    
    if request.method == "POST":
        # Get form data
        lname = request.form.get("Lname", "").strip()
        fname = request.form.get("Fname", "").strip()
        mname = request.form.get("Mname", "").strip()
        address = request.form.get("address", "").strip()
        email = request.form.get("email", "").strip()
        contact = request.form.get("contact", "").strip()
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user_type = request.form.get("user_type", "")
        
        # Validation
        if not lname or not fname or not username or not password or not user_type:
            error = "Required fields marked with * are mandatory."
        elif len(password) < 6:
            error = "Password must be at least 6 characters."
        elif user_type not in ["ADMIN", "ALUMNI COORDINATOR"]:
            error = "Invalid user type selected."
        else:
            # Convert uppercase user_type to proper case for database consistency
            if user_type == "ADMIN":
                user_type = "Admin"
            elif user_type == "ALUMNI COORDINATOR":
                user_type = "Alumni Coordinator"
                
            db = get_db()
            cursor = db.cursor()
            try:
                # Check if username already exists
                cursor.execute("SELECT id FROM user_login WHERE username = %s", (username,))
                if cursor.fetchone():
                    error = "Username already exists. Please choose another."
                else:
                    # Insert into user_information
                    cursor.execute("""
                        INSERT INTO user_information
                        (Lname, Fname, Mname, address, email, contact, addedBy, dateAdded)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (
                        lname, fname, mname, address, email, 
                        int(contact) if contact.isdigit() else 0,
                        "Self-Registration", date.today()
                    ))
                    
                    user_id = cursor.lastrowid
                    
                    # Insert into user_login with plain text password
                    cursor.execute("""
                        INSERT INTO user_login
                        (id, username, password, user_type)
                        VALUES (%s,%s,%s,%s)
                    """, (user_id, username, password, user_type))
                    
                    db.commit()
                    success = "Account created successfully! You can now log in."
                    
            except Exception as e:
                import traceback
                print("ERROR - Save user public failed:", str(e))
                traceback.print_exc()
                db.rollback()
                error = f"Registration failed: {str(e)}"
            finally:
                cursor.close()
                db.close()
    
    return render_template("createuser.html", user_type=None, users=None, error=error, success=success)

@app.route("/save-user", methods=["POST"])
@admin_required
def save_user():

    # Get and validate user_type
    user_type = request.form.get("user_type", "").strip()
    
    # Validate user_type
    if user_type not in ["ADMIN", "ALUMNI COORDINATOR"]:
        return "Invalid user type selected.", 400
    
    # Convert to proper case for database consistency
    if user_type == "ADMIN":
        user_type = "Admin"
    elif user_type == "ALUMNI COORDINATOR":
        user_type = "Alumni Coordinator"

    db = get_db()
    cursor = db.cursor()

    try:
        print("FORM DATA:", request.form)

        # ================= USER INFORMATION =================
        cursor.execute("""
            INSERT INTO user_information
            (Lname, Fname, Mname, address, email, contact, addedBy, dateAdded)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            request.form.get("Lname"),
            request.form.get("Fname"),
            request.form.get("Mname"),
            request.form.get("address"),
            request.form.get("email"),
            int(request.form.get("contact") or 0),
            session.get("username"),
            date.today()
        ))

        user_id = cursor.lastrowid

        # ================= USER LOGIN =================
        cursor.execute("""
            INSERT INTO user_login
            (id, username, password, user_type)
            VALUES (%s,%s,%s,%s)
        """, (
            user_id,
            request.form.get("username"),
            request.form.get("password"),
            user_type
        ))

        db.commit()

        log_activity(f"Created user {request.form.get('username')}")

        return redirect("/create-user")

    except Exception as e:
        import traceback
        print("ERROR - Save user failed:", str(e))
        traceback.print_exc()
        db.rollback()

        return render_template(
            "createuser.html",
            error=str(e),
            user_type=session.get("user_type")
        )

    finally:
        cursor.close()
        db.close()


@app.route("/get-user/<int:id>")
@admin_required
def get_user(id):

    db = get_db()
    cursor = db.cursor()

    # USER INFO
    cursor.execute("""
        SELECT address, email, contact, addedBy, dateAdded
        FROM user_information
        WHERE id = %s
    """, (id,))
    info = cursor.fetchone()

    # LOGIN INFO
    cursor.execute("""
        SELECT username, password, user_type
        FROM user_login
        WHERE id = %s
    """, (id,))
    login = cursor.fetchone()

    cursor.close()
    db.close()

    return jsonify({
        "info": info or {},
        "login": login or {}
    })


def resequence_alumni_ids():
    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute("SET FOREIGN_KEY_CHECKS = 0")

        cursor.execute("SELECT alumni_id FROM alumni_table ORDER BY alumni_id")
        rows = cursor.fetchall()

        new_id = 1
        for (old_id,) in rows:
            if old_id != new_id:

                # update main table
                cursor.execute("""
                    UPDATE alumni_table
                    SET alumni_id = %s
                    WHERE alumni_id = %s
                """, (new_id, old_id))

                # update degree
                cursor.execute("""
                    UPDATE alumni_degree
                    SET alumni_id = %s
                    WHERE alumni_id = %s
                """, (new_id, old_id))

                # update employment
                cursor.execute("""
                    UPDATE alumni_employment
                    SET alumni_id = %s
                    WHERE alumni_id = %s
                """, (new_id, old_id))

            new_id += 1

        cursor.execute("ALTER TABLE alumni_table AUTO_INCREMENT = 1")

        cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
        db.commit()

    except Exception as e:
        db.rollback()
        print("Resequence error:", e)

    finally:
        cursor.close()
        db.close()


@app.route("/export-records")
@login_required
def export_records():
    search = request.args.get("search", "").strip()
    format_type = request.args.get("format", "csv")

    db = get_db()
    cursor = db.cursor()

    # reuse the exact same search logic — no pagination
    records_query, _, params, _ = build_search_query(search)

    cursor.execute(records_query, tuple(params))
    data = cursor.fetchall()

    cursor.close()
    db.close()

    def safe(v):
        if v is None:
            return ""
        if hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
        return str(v)

    headers = [
        "Student No", "Last Name", "First Name", "Middle Name",
        "Address", "Email", "Contact",
        "Program", "Major", "Admission Date", "Graduation Date",
        "Employment Status", "Job Title", "Sector", "Degree Relevance",
        "Added By", "Date Added"
    ]

    def row_values(r):
        return [
            safe(r[1]), safe(r[2]), safe(r[3]), safe(r[4]),
            safe(r[5]), safe(r[6]), safe(r[7]),
            safe(r[10]), safe(r[11]), safe(r[12]), safe(r[13]),
            safe(r[14]), safe(r[15]), safe(r[16]), safe(r[17]),
            safe(r[8]), safe(r[9])
        ]

    # ================= CSV =================
    if format_type == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for r in data:
            writer.writerow(row_values(r))
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode()),
            mimetype="text/csv",
            as_attachment=True,
            download_name="alumni_records.csv"
        )

    # ================= EXCEL =================
    elif format_type == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Alumni Records"
        ws.append(headers)
        for r in data:
            ws.append(row_values(r))
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name="alumni_records.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # ================= PDF =================
    elif format_type == "pdf":
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=LETTER,
            rightMargin=50, leftMargin=50,
            topMargin=60, bottomMargin=60
        )
        elements = []
        styles = getSampleStyleSheet()
        normal = ParagraphStyle(name="exp_normal", fontName="Times-Roman", fontSize=9, leading=12)
        bold_st = ParagraphStyle(name="exp_bold", fontName="Times-Bold", fontSize=9)
        header_st = ParagraphStyle(name="exp_header", fontName="Times-Bold", fontSize=11, leading=14)

        try:
            logo = Image("static/images/PNG (transparent background).png", 50, 50)
        except Exception:
            logo = Paragraph("", normal)

        header_tbl = Table([
            [logo, Paragraph(
                "Nueva Ecija University of Science and Technology<br/>"
                "Talavera Academic Extension Campus<br/><br/>"
                "<b>Alumni Records Export</b>",
                header_st
            )]
        ], colWidths=[60, 440])
        header_tbl.setStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")])
        elements.append(header_tbl)
        elements.append(Spacer(1, 6))

        if search:
            elements.append(Paragraph(f"<b>Search Filter:</b> {search}", normal))
            elements.append(Spacer(1, 4))

        elements.append(Paragraph(f"Total Records: {len(data)}", bold_st))
        elements.append(Spacer(1, 10))

        col_headers = [
            "Stud No", "Last Name", "First Name",
            "Program", "Major", "Graduation",
            "Email", "Contact", "Status"
        ]
        col_widths = [65, 65, 60, 60, 60, 55, 80, 65, 55]
        table_data = [col_headers]

        for r in data:
            table_data.append([
                safe(r[1]), safe(r[2]), safe(r[3]),
                safe(r[10]), safe(r[11]), safe(r[13]),
                safe(r[6]), safe(r[7]), safe(r[14])
            ])

        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a73e8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Times-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("FONTNAME", (0, 1), (-1, -1), "Times-Roman"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4ff")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(tbl)

        doc.build(elements, onFirstPage=draw_layout, onLaterPages=draw_layout)
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name="alumni_records.pdf",
            mimetype="application/pdf"
        )

    return "Invalid format", 400



@app.template_filter('studnum')
def studnum_filter(value):
    if not value:
        return "-"
    return value if value.startswith("TAL-") else f"TAL-{value}"



def resequence_alumni_ids():
    db = get_db()
    cursor = db.cursor()

    try:

        cursor.execute("SET FOREIGN_KEY_CHECKS = 0")

        cursor.execute("SELECT alumni_id FROM alumni_table ORDER BY alumni_id ASC")
        rows = cursor.fetchall()

        new_id = 1

        for (old_id,) in rows:

            if old_id != new_id:

                temp_id = -(old_id + 1000000)  # safe unique temp id

                # STEP 1: move to temp ID
                cursor.execute("""
                    UPDATE alumni_table
                    SET alumni_id = %s
                    WHERE alumni_id = %s
                """, (temp_id, old_id))

                cursor.execute("""
                    UPDATE alumni_degree
                    SET alumni_id = %s
                    WHERE alumni_id = %s
                """, (temp_id, old_id))

                cursor.execute("""
                    UPDATE alumni_employment
                    SET alumni_id = %s
                    WHERE alumni_id = %s
                """, (temp_id, old_id))

                # STEP 2: assign final correct ID
                cursor.execute("""
                    UPDATE alumni_table
                    SET alumni_id = %s
                    WHERE alumni_id = %s
                """, (new_id, temp_id))

                cursor.execute("""
                    UPDATE alumni_degree
                    SET alumni_id = %s
                    WHERE alumni_id = %s
                """, (new_id, temp_id))

                cursor.execute("""
                    UPDATE alumni_employment
                    SET alumni_id = %s
                    WHERE alumni_id = %s
                """, (new_id, temp_id))

            new_id += 1

        cursor.execute("ALTER TABLE alumni_table AUTO_INCREMENT = 1")

        cursor.execute("SET FOREIGN_KEY_CHECKS = 1")

        db.commit()

    except Exception as e:
        db.rollback()
        print("Resequence error:", e)

    finally:
        cursor.close()
        db.close()



def format_stud_num(value):
    if not value:
        return "-"
    value = str(value)
    return value if value.startswith("TAL-") else f"TAL-{value}"



def safe(v):
    """Safe value conversion for database storage"""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v)

def safe_date(v):
    """Safe date conversion for database storage"""
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v)


def log_activity(activity):
    if "username" not in session:
        return

    db = get_db()
    cursor = db.cursor()

    try:
        # Convert user_type to uppercase to match activity_logs table enum
        user_type = session.get("user_type", "")
        if user_type == "Admin":
            db_user_type = "ADMIN"
        elif user_type == "Alumni Coordinator":
            db_user_type = "ALUMNI COORDINATOR"
        else:
            db_user_type = user_type  # Use as-is for other types
        
        cursor.execute("""
            INSERT INTO activity_logs (username, user_type, activity, date_time)
            VALUES (%s, %s, %s, %s)
        """, (
            session.get("username"),
            db_user_type,
            activity,
            datetime.now()
        ))

        db.commit()

    except Exception as e:
        print("Activity log error:", e)
        db.rollback()

    finally:
        cursor.close()
        db.close()





# ================================================================
# IDAGDAG SA app.py — BAGO ang `if __name__ == "__main__":` line
# I-copy paste ang dalawang routes na ito
# ================================================================

# ================= HELPER: Build filtered query =================
def build_filtered_query(args):
    """
    Builds a parameterized SQL query based on advanced filter args.
    Returns (query_string, params_list)
    """

    date_from  = args.get("date_from", "").strip()
    date_to    = args.get("date_to", "").strip()
    grad_from  = args.get("grad_from", "").strip()
    grad_to    = args.get("grad_to", "").strip()
    program    = args.get("program", "").strip()
    major      = args.get("major", "").strip()
    lastname   = args.get("lastname", "").strip()
    firstname  = args.get("firstname", "").strip()
    address    = args.get("address", "").strip()
    email      = args.get("email", "").strip()
    contact    = args.get("contact", "").strip()

    # clean contact (digits only)
    contact_clean = re.sub(r"\D", "", contact)

    query = """
        SELECT
            a.alumni_id, a.stud_num, a.last_name, a.first_name, a.middle_name,
            a.address, a.email, a.contact_num,
            d.program, d.major, d.graduation_date,
            e.employment_status, e.employment_sector, e.job_title, e.degree_relevance_to_work,
            a.added_by, a.date_added
        FROM alumni_table a
        LEFT JOIN alumni_degree d ON a.alumni_id = d.alumni_id
        LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
        WHERE 1=1
    """

    params = []

    # ---- date_added range ----
    if date_from:
        query += " AND DATE(a.date_added) >= %s"
        params.append(date_from)

    if date_to:
        query += " AND DATE(a.date_added) <= %s"
        params.append(date_to)

    # ---- graduation date range ----
    if grad_from:
        query += " AND DATE(d.graduation_date) >= %s"
        params.append(grad_from)

    if grad_to:
        query += " AND DATE(d.graduation_date) <= %s"
        params.append(grad_to)

    # ---- program (partial match) ----
    if program:
        query += " AND LOWER(COALESCE(d.program,'')) LIKE %s"
        params.append(f"%{program.lower()}%")

    # ---- major (partial match) ----
    if major:
        query += " AND LOWER(COALESCE(d.major,'')) LIKE %s"
        params.append(f"%{major.lower()}%")

    # ---- last name (partial match) ----
    if lastname:
        query += " AND LOWER(COALESCE(a.last_name,'')) LIKE %s"
        params.append(f"%{lastname.lower()}%")

    # ---- first name (partial match) ----
    if firstname:
        query += " AND LOWER(COALESCE(a.first_name,'')) LIKE %s"
        params.append(f"%{firstname.lower()}%")

    # ---- address (partial match, handles partial barangay/town names) ----
    if address:
        query += " AND LOWER(COALESCE(a.address,'')) LIKE %s"
        params.append(f"%{address.lower()}%")

    # ---- email (partial match) ----
    if email:
        query += " AND LOWER(COALESCE(a.email,'')) LIKE %s"
        params.append(f"%{email.lower()}%")

    # ---- contact (digits-only partial match) ----
    if contact_clean:
        query += """
            AND REPLACE(REPLACE(COALESCE(CAST(a.contact_num AS CHAR),''),' ',''),'-','') LIKE %s
        """
        params.append(f"%{contact_clean}%")

    query += " GROUP BY a.alumni_id ORDER BY a.alumni_id ASC"

    return query, params


# ================= ROUTE: Export Filtered Records =================
@app.route("/export-filtered")
@login_required
def export_filtered():

    format_type = request.args.get("format", "excel")

    db = get_db()
    cursor = db.cursor()

    query, params = build_filtered_query(request.args)
    cursor.execute(query, tuple(params))
    data = cursor.fetchall()

    cursor.close()
    db.close()

    def safe(v):
        return v if v is not None else ""

    # ================= EXCEL =================
    if format_type == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Alumni Records"

        ws.append([
            "Student No", "Last Name", "First Name", "Middle Name",
            "Address", "Email", "Contact",
            "Program", "Major", "Admission Date", "Graduation Date",
            "Employment Status", "Sector", "Job Title", "Degree Relevance",
            "Added By", "Date Added"
        ])

        for r in data:
            ws.append([
                safe(r[1]),  safe(r[2]),  safe(r[3]),  safe(r[4]),
                safe(r[5]),  safe(r[6]),  safe(r[7]),
                safe(r[8]),  safe(r[9]),  safe(r[10]), safe(r[11]),
                safe(r[12]), safe(r[13]), safe(r[14]), safe(r[15]),
                safe(r[16]), safe(r[17])
            ])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name="alumni_filtered.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # ================= PDF =================
    elif format_type == "pdf":
        buffer = BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=LETTER,
            rightMargin=50,
            leftMargin=50,
            topMargin=60,
            bottomMargin=60
        )

        elements = []
        styles = getSampleStyleSheet()

        normal    = ParagraphStyle(name="normal_f", fontName="Times-Roman",  fontSize=9,  leading=12)
        bold_st   = ParagraphStyle(name="bold_f",   fontName="Times-Bold",   fontSize=9)
        header_st = ParagraphStyle(name="header_f", fontName="Times-Bold",   fontSize=11, leading=14)
        title_st  = ParagraphStyle(name="title_f",  fontName="Times-Bold",   fontSize=13, leading=16, alignment=1)

        # ---- try logo ----
        try:
            logo = Image("static/images/PNG (transparent background).png", 50, 50)
        except Exception:
            logo = Paragraph("", normal)

        # ---- page header ----
        header_tbl = Table([
            [
                logo,
                Paragraph(
                    "Nueva Ecija University of Science and Technology<br/>"
                    "Talavera Academic Extension Campus<br/><br/>"
                    "<b>Filtered Alumni Records</b>",
                    header_st
                )
            ]
        ], colWidths=[60, 440])

        header_tbl.setStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])

        elements.append(header_tbl)
        elements.append(Spacer(1, 6))

        # ---- filter summary ----
        filter_lines = []
        field_labels = {
            "date_from"  : "Date Added From",
            "date_to"    : "Date Added To",
            "grad_from"  : "Graduation From",
            "grad_to"    : "Graduation To",
            "program"    : "Program",
            "major"      : "Major",
            "lastname"   : "Last Name",
            "firstname"  : "First Name",
            "address"    : "Address",
            "email"      : "Email",
            "contact"    : "Contact"
        }

        for key, label in field_labels.items():
            val = request.args.get(key, "").strip()
            if val:
                filter_lines.append(f"<b>{label}:</b> {val}")

        if filter_lines:
            elements.append(Paragraph("Applied Filters: " + " | ".join(filter_lines), normal))
            elements.append(Spacer(1, 4))

        elements.append(Paragraph(f"Total Records: {len(data)}", bold_st))
        elements.append(Spacer(1, 10))

        # ---- table header ----
        col_headers = [
            "Stud No", "Last Name", "First Name",
            "Program", "Major", "Graduation",
            "Email", "Contact", "Status"
        ]

        col_widths = [65, 65, 60, 60, 60, 55, 80, 65, 55]

        table_data = [col_headers]

        for r in data:
            grad = safe(r[11])
            if hasattr(grad, "strftime"):
                grad = grad.strftime("%Y-%m-%d")

            table_data.append([
                safe(r[1]),
                safe(r[2]),
                safe(r[3]),
                safe(r[8]),
                safe(r[9]),
                str(grad),
                safe(r[6]),
                safe(r[7]),
                safe(r[12])
            ])

        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)

        tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#1a73e8")),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",    (0, 0), (-1, 0),  "Times-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 7.5),
            ("FONTNAME",    (0, 1), (-1, -1), "Times-Roman"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4ff")]),
            ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ]))

        elements.append(tbl)

        doc.build(elements, onFirstPage=draw_layout, onLaterPages=draw_layout)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name="alumni_filtered.pdf",
            mimetype="application/pdf"
        )

    return "Invalid format", 400


# ================= ROUTE: Count for preview =================
@app.route("/export-filtered-count")
@login_required
def export_filtered_count():

    db = get_db()
    cursor = db.cursor()

    # reuse the same helper, pero SELECT COUNT lang
    _, params = build_filtered_query(request.args)

    count_query = """
        SELECT COUNT(DISTINCT a.alumni_id)
        FROM alumni_table a
        LEFT JOIN alumni_degree d ON a.alumni_id = d.alumni_id
        LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
        WHERE 1=1
    """

    date_from  = request.args.get("date_from", "").strip()
    date_to    = request.args.get("date_to", "").strip()
    grad_from  = request.args.get("grad_from", "").strip()
    grad_to    = request.args.get("grad_to", "").strip()
    program    = request.args.get("program", "").strip()
    major      = request.args.get("major", "").strip()
    lastname   = request.args.get("lastname", "").strip()
    firstname  = request.args.get("firstname", "").strip()
    address    = request.args.get("address", "").strip()
    email      = request.args.get("email", "").strip()
    contact    = request.args.get("contact", "").strip()
    contact_clean = re.sub(r"\D", "", contact)

    count_params = []

    if date_from:
        count_query += " AND DATE(a.date_added) >= %s"
        count_params.append(date_from)

    if date_to:
        count_query += " AND DATE(a.date_added) <= %s"
        count_params.append(date_to)

    if grad_from:
        count_query += " AND DATE(d.graduation_date) >= %s"
        count_params.append(grad_from)

    if grad_to:
        count_query += " AND DATE(d.graduation_date) <= %s"
        count_params.append(grad_to)

    if program:
        count_query += " AND LOWER(COALESCE(d.program,'')) LIKE %s"
        count_params.append(f"%{program.lower()}%")

    if major:
        count_query += " AND LOWER(COALESCE(d.major,'')) LIKE %s"
        count_params.append(f"%{major.lower()}%")

    if lastname:
        count_query += " AND LOWER(COALESCE(a.last_name,'')) LIKE %s"
        count_params.append(f"%{lastname.lower()}%")

    if firstname:
        count_query += " AND LOWER(COALESCE(a.first_name,'')) LIKE %s"
        count_params.append(f"%{firstname.lower()}%")

    if address:
        count_query += " AND LOWER(COALESCE(a.address,'')) LIKE %s"
        count_params.append(f"%{address.lower()}%")

    if email:
        count_query += " AND LOWER(COALESCE(a.email,'')) LIKE %s"
        count_params.append(f"%{email.lower()}%")

    if contact_clean:
        count_query += """
            AND REPLACE(REPLACE(COALESCE(CAST(a.contact_num AS CHAR),''),' ',''),'-','') LIKE %s
        """
        count_params.append(f"%{contact_clean}%")

    cursor.execute(count_query, tuple(count_params))
    result = cursor.fetchone()
    count = result['COUNT(*)'] if result else 0

    cursor.close()
    db.close()

    return jsonify({"count": count})


def build_search_query(search, per_page=None, offset=None):
    """
    Unified search query used by BOTH /records (display) and /export-records (download).
    Returns (records_query, count_query, params, count_params)
    Pagination params are appended only when per_page/offset are given.
    """

    base_select = """
        SELECT
            a.alumni_id, a.stud_num, a.photo, a.last_name, a.first_name, a.middle_name,
            a.address, a.email, a.contact_num, a.added_by, a.date_added,
            d.program, d.major, d.graduation_date,
            e.employment_status, e.job_title, e.employment_sector, e.degree_relevance_to_work
        FROM alumni_table a
        LEFT JOIN alumni_degree d ON a.alumni_id = d.alumni_id
        LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
    """

    count_select = """
        SELECT COUNT(DISTINCT a.alumni_id)
        FROM alumni_table a
        LEFT JOIN alumni_degree d ON a.alumni_id = d.alumni_id
        LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
    """

    if not search:
        where = ""
        params = []
    else:
        search_lower = search.lower().strip()
        like = f"%{search_lower}%"
        search_contact = re.sub(r"\D", "", search)
        is_year = search.isdigit() and len(search) == 4

        if is_year:
            where = """
                WHERE (
                    YEAR(a.date_added)        = %s OR
                    YEAR(d.graduation_date)   = %s
                )
            """
            params = [search, search]
        else:
            where = """
                WHERE (
                    LOWER(COALESCE(a.stud_num,          '')) LIKE %s OR
                    LOWER(COALESCE(a.first_name,        '')) LIKE %s OR
                    LOWER(COALESCE(a.middle_name,       '')) LIKE %s OR
                    LOWER(COALESCE(a.last_name,         '')) LIKE %s OR
                    LOWER(COALESCE(a.email,             '')) LIKE %s OR
                    LOWER(COALESCE(a.address,           '')) LIKE %s OR
                    LOWER(COALESCE(d.program,           '')) LIKE %s OR
                    LOWER(COALESCE(d.major,             '')) LIKE %s OR
                    LOWER(COALESCE(e.job_title,         '')) LIKE %s OR
                    LOWER(COALESCE(e.employment_status, '')) LIKE %s OR
                    LOWER(COALESCE(e.employment_sector, '')) LIKE %s OR
                    REPLACE(REPLACE(COALESCE(CAST(a.contact_num AS CHAR),''),' ',''),'-','') LIKE %s OR
                    CAST(d.graduation_date   AS CHAR) LIKE %s OR
                    LOWER(COALESCE(e.degree_relevance_to_work, '')) LIKE %s
                )
            """
            contact_like = f"%{search_contact}%" if search_contact else like
            params = [
                like, like, like, like, like,  # stud_num, first, middle, last, email
                like,  # address
                like, like,  # program, major
                like, like, like,  # job_title, emp_status, emp_sector
                contact_like,  # contact (digits-only match)
                like, like,  # graduation_date, admission_date
                like  # degree_relevance_to_work
            ]

    group_order = " GROUP BY a.alumni_id ORDER BY a.alumni_id ASC"

    records_query = base_select + where + group_order
    count_query = count_select + where

    # attach pagination to records query only
    paginated_params = list(params)
    if per_page is not None and offset is not None:
        records_query += " LIMIT %s OFFSET %s"
        paginated_params += [per_page, offset]

    return records_query, count_query, paginated_params, list(params)












# ================= API: Employment Status Summary =================
@app.route("/api/employment_summary")
@login_required
def employment_summary():
    db = get_db()
    cursor = db.cursor()

    cursor.execute("""
        SELECT
            COALESCE(NULLIF(TRIM(e.employment_status), ''), 'Unknown') AS status,
            COUNT(DISTINCT a.alumni_id) AS total
        FROM alumni_table a
        LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
        GROUP BY status
        ORDER BY total DESC
    """)

    rows = cursor.fetchall()
    cursor.close()
    db.close()

    labels = [r['label'] for r in rows]
    counts = [r['count'] for r in rows]

    return jsonify({"labels": labels, "counts": counts})


# ================= API: Degree Relevance Summary =================
@app.route("/api/relevance_summary")
@login_required
def relevance_summary():
    db = get_db()
    cursor = db.cursor()

    cursor.execute("""
        SELECT
            COALESCE(NULLIF(TRIM(e.degree_relevance_to_work), ''), 'Unknown') AS relevance,
            COUNT(DISTINCT a.alumni_id) AS total
        FROM alumni_table a
        LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
        WHERE e.employment_status IS NOT NULL
          AND TRIM(e.employment_status) != ''
          AND LOWER(TRIM(e.employment_status)) != 'unemployed'
        GROUP BY relevance
        ORDER BY total DESC
    """)

    rows = cursor.fetchall()
    cursor.close()
    db.close()

    labels = [r['label'] for r in rows]
    counts = [r['count'] for r in rows]

    return jsonify({"labels": labels, "counts": counts})


@app.route("/register", methods=["GET", "POST"])
def register():
    error = None
    success = None

    if request.method == "POST":
        step = request.form.get("step", "create")

        if step == "create":
            alumni_id = request.form.get("alumni_id")
            lname = request.form.get("last_name", "").strip()
            fname = request.form.get("first_name", "").strip()
            mname = request.form.get("middle_name", "").strip()
            email = request.form.get("email", "").strip()
            address = request.form.get("address", "").strip()
            password = request.form.get("password", "")
            confirm = request.form.get("confirm_password", "")

            # Handle photo upload
            photo_file = request.files.get("photo")
            photo_filename = None
            
            if photo_file and photo_file.filename:
                # Validate file type
                allowed_extensions = {'jpg', 'jpeg', 'png'}
                file_ext = photo_file.filename.rsplit('.', 1)[1].lower() if '.' in photo_file.filename else ''
                
                if file_ext not in allowed_extensions:
                    error = "Invalid file type. Please upload JPG or PNG image."
                elif photo_file.content_length > 5 * 1024 * 1024:  # 5MB limit
                    error = "File too large. Please upload an image smaller than 5MB."
                else:
                    # Generate unique filename
                    filename = secure_filename(photo_file.filename)
                    unique_filename = f"{uuid.uuid4().hex}_{filename}"
                    photo_path = os.path.join('static', 'uploads', 'profile_photos', unique_filename)
                    
                    # Save file
                    os.makedirs(os.path.dirname(photo_path), exist_ok=True)
                    photo_file.save(photo_path)
                    photo_filename = f"uploads/profile_photos/{unique_filename}"

            if not email:
                error = "Email is required."
            elif not password or len(password) < 6:
                error = "Password must be at least 6 characters."
            elif password != confirm:
                error = "Passwords do not match."
            elif not address or len(address.strip()) == 0:
                error = "Address is required."
            elif not photo_filename:
                error = "Profile photo is required."
            else:
                db = get_db()
                cursor = db.cursor()
                try:
                    # Check email uniqueness in alumni_table
                    cursor.execute(
                        "SELECT alumni_id FROM alumni_table WHERE email=%s",
                        (email,)
                    )
                    if cursor.fetchone():
                        error = "An account with this email already exists."
                    else:
                        # Generate student number
                        stud_num = f"TAL-{datetime.now().year}-{str(int(time.time()) % 100000).zfill(5)}"
                        
                        cursor.execute("""
                            INSERT INTO alumni_table
                            (stud_num, last_name, first_name, middle_name, email, alumni_password, address, photo, added_by, date_added)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (stud_num, lname, fname, mname, email, password, address, photo_filename, 'alumni_self_service', date.today()))
                        
                        db.commit()
                        success = "Account created successfully! You can now log in as Alumni."
                except Exception as e:
                    import traceback
                    print("ERROR - Registration failed:", str(e))
                    traceback.print_exc()
                    db.rollback()
                    error = str(e)
                finally:
                    cursor.close()
                    db.close()

    return render_template(
        "register.html",
        step="create",
        error=error,
        success=success,
        alumni=None
    )


@app.route("/my-profile", methods=["GET", "POST"])
@login_required
def my_profile():
    """
    ALUMNI users can only view and edit their own record.
    """
    if session.get("user_type") != "Alumni":
        return redirect("/dashboard")

    db = get_db()
    cursor = db.cursor()
    try:
        # Fetch programs for dropdown
        cursor.execute("SELECT program_name FROM programs ORDER BY program_name")
        programs = [row['program_name'] for row in cursor.fetchall()]
        
        # Get alumni info using alumni_id from session
        cursor.execute("SELECT * FROM alumni_table WHERE alumni_id=%s", (session.get("alumni_id"),))
        alumni = cursor.fetchone()

    except Exception as e:
        cursor.close()
        db.close()
        return render_template(
            "my_profile.html",
            alumni=None,
            degree=None,
            employment=None,
            error="Failed to fetch data: " + str(e),
            user_type="ALUMNI"
        )

    if not alumni:
        cursor.close()
        db.close()
        return render_template(
            "my_profile.html",
            alumni=None,
            degree=None,
            employment=None,
            error="Account not found. Please contact the coordinator.",
            user_type="ALUMNI"
        )

    # Initialize variables
    alumni = None
    educ = None
    employ = []
    error = None
    success = None
    can_edit = False  # Flag to control edit mode access
    
    # Fetch existing alumni record if it exists - prioritize records with employment data
    cursor.execute("""
        SELECT at.* FROM alumni_table at
        LEFT JOIN alumni_employment e ON at.alumni_id = e.alumni_id
        WHERE at.alumni_id = %s 
        ORDER BY 
            CASE WHEN e.employment_status = 'Employed' AND e.job_title IS NOT NULL THEN 0 ELSE 1 END,
            at.alumni_id DESC
        LIMIT 1
    """, (session.get("alumni_id"),))
    existing_alumni = cursor.fetchone()
    
    if existing_alumni:
        # Use existing alumni record
        alumni = existing_alumni
        alumni['has_record'] = True
        
        # Fetch education and employment data
        cursor.execute("SELECT * FROM alumni_degree WHERE alumni_id=%s LIMIT 1", (alumni['alumni_id'],))
        educ = cursor.fetchone()
        
        cursor.execute("""
            SELECT * FROM alumni_employment 
            WHERE alumni_id=%s 
            ORDER BY employment_id ASC
        """, (alumni['alumni_id'],))
        employ = cursor.fetchall()
    else:
        # Get basic alumni info from session for new profiles
        cursor.execute("SELECT first_name, last_name, middle_name, email FROM alumni_table WHERE alumni_id=%s", (session.get("alumni_id"),))
        basic_alumni = cursor.fetchone()
        
        # Create blank profile from session alumni info
        alumni = {
            'alumni_id': session.get("alumni_id"),
            'last_name': basic_alumni['last_name'] if basic_alumni else '',
            'first_name': basic_alumni['first_name'] if basic_alumni else '',
            'middle_name': basic_alumni.get('middle_name', '') if basic_alumni else '',
            'email': basic_alumni['email'] if basic_alumni else '',
            'address': '',
            'contact_num': '',
            'photo': None,
            'date_added': '',
            'has_record': False  # Mark that this alumni needs to create their record
        }
        
        # Initialize educ and employ as empty for new profiles
        educ = None
        employ = []
    
    # Check for success parameter from URL AFTER fetching fresh data
    if request.args.get('success') == '1':
        success = "Record Successfully Updated"
    
    # Check if alumni has approved update request
    if alumni and alumni.get('has_record', False):
        # For existing alumni records, check alumni_table ID
        cursor.execute("""
            SELECT id FROM alumni_update_requests 
            WHERE alumni_id = %s AND status = 'approved'
            ORDER BY requested_at DESC LIMIT 1
        """, (alumni['alumni_id'],))
        approved_request = cursor.fetchone()
        can_edit = approved_request is not None

    if request.method == "POST":
        try:
            # Check if this is a new record creation
            is_new_record = not alumni.get('has_record', False)
            
            # ---------- VALIDATION ----------
            if is_new_record:
                # Student Number validation
                stud_num = request.form.get("stud_num", "").strip()
                if not stud_num:
                    error = "Student number is required."
                else:
                    # Remove TAL- prefix if present and validate format
                    clean_stud_num = stud_num.replace("TAL-", "").replace(" ", "")
                    if not re.match(r"^\d{4}-\d{5}$", clean_stud_num):
                        error = "Student number must be in format YYYY-XXXXX (e.g., 2023-00238)."
                    else:
                        # Auto-prepend TAL- for database storage
                        stud_num = f"TAL-{clean_stud_num}"
                if not error:
                    # Check for duplicate student number
                    cursor.execute("SELECT alumni_id FROM alumni_table WHERE stud_num=%s", (stud_num,))
                    if cursor.fetchone():
                        error = "This student number already exists. Please use your unique student number."
                
                # Email validation
                email = request.form.get("email", "").strip()
                if not email:
                    error = "Email is required." if not error else error
                elif not re.match(r"^[^@]+@[^@]+\.[^@]+$", email):
                    error = "Please enter a valid email address." if not error else error
            
            # ---------- Contact ----------
            contact = re.sub(r"\D", "", request.form.get("contact_num", ""))
            if not contact or len(contact) != 11 or not contact.startswith("09"):
                error = "Contact number must be 11 digits starting with 09." if not error else error
            else:
                # ---------- Photo ----------
                old_photo = alumni.get("photo") if alumni else None
                remove_flag = request.form.get("remove_photo", "0")
                photo_file = request.files.get("photo")
                new_photo = old_photo

                if remove_flag == "1":
                    delete_photo(old_photo)
                    new_photo = None
                elif photo_file and photo_file.filename:
                    saved = save_photo(photo_file)
                    if saved:
                        delete_photo(old_photo)
                        new_photo = saved

                if is_new_record:
                    # ---------- CREATE NEW ALUMNI RECORD ----------
                    # Get student number from form or generate if not provided
                    stud_num = request.form.get("stud_num", "").strip()
                    if not stud_num:
                        stud_num = f"TAL-{datetime.now().year}-{str(session.get('alumni_id')).zfill(5)}"
                    
                    cursor.execute("""
                        INSERT INTO alumni_table
                        (stud_num, last_name, middle_name, first_name, address, email, contact_num, photo, added_by, date_added)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (
                        stud_num,
                        alumni['last_name'],
                        alumni.get('middle_name', ''),
                        alumni['first_name'],
                        request.form.get("address"),
                        alumni['email'],
                        contact,
                        new_photo,
                        "alumni_self_service",
                        date.today()
                    ))
                    alumni_id = cursor.lastrowid
                    
                    # ---------- DEGREE ----------
                    program = request.form.get("program", "")
                    major = request.form.get("major", "")
                    grad_date = request.form.get("graduation_date") or None
                    
                    # Validate Graduation Date
                    if grad_date:
                        try:
                            grad_date_obj = datetime.strptime(grad_date, '%Y-%m-%d').date()
                            if grad_date_obj > date.today():
                                error = "Graduation Date cannot be greater than current date."
                                return render_template("my_profile.html", alumni=alumni, degree=educ, employment=employ, error=error, user_type="Alumni", now=datetime.now())
                        except ValueError:
                            error = "Invalid Graduation Date format."
                            return render_template("my_profile.html", alumni=alumni, degree=educ, employment=employ, error=error, user_type="Alumni", now=datetime.now())
                    
                    if program:
                        cursor.execute("""
                            INSERT INTO alumni_degree
                            (alumni_id, program, major, graduation_date, added_by, date_added)
                            VALUES (%s,%s,%s,%s,%s,%s)
                        """, (
                            alumni_id,
                            program,
                            major,
                            grad_date,
                            "alumni_self_service",
                            date.today()
                        ))
                    
                    # ---------- EMPLOYMENT ----------
                    employment_status = request.form.get("employment_status", "")
                    if employment_status:
                        if employment_status == "Unemployed":
                            cursor.execute("""
                                INSERT INTO alumni_employment
                                (alumni_id, employment_status, employment_sector, job_title, degree_relevance_to_work, added_by, date_added)
                                VALUES (%s,%s,%s,%s,%s,%s,%s)
                            """, (alumni_id, "Unemployed", None, None, None, "alumni_self_service", date.today()))
                        else:
                            job_titles = request.form.getlist("job_title[]")
                            sectors = request.form.getlist("employment_sector[]")
                            relevances = request.form.getlist("degree_relevance_to_work[]")
                            
                            if job_titles:
                                for i, job_title in enumerate(job_titles):
                                    sector = sectors[i] if i < len(sectors) else ""
                                    relevance = relevances[i] if i < len(relevances) else ""
                                    
                                    # Handle "Let AI Decide" option
                                    if relevance == "Let AI Decide":
                                        # Call AI scan relevance function
                                        try:
                                            program = request.form.get("program", "")
                                            major = request.form.get("major", "")
                                            
                                            # Prepare AI scan request
                                            ai_data = {
                                                "job_title": job_title,
                                                "program": program,
                                                "major": major
                                            }
                                            
                                            # Import the scan function logic here
                                            prompt = (
                                                f"A college graduate completed a {program} degree "
                                                f"with a major in {major}. "
                                                f"Their current job title is: '{job_title}'.\n\n"
                                                f"Based on this, classify how relevant their degree is to their work. "
                                                f"Reply with EXACTLY one of these four options and nothing else:\n"
                                                f"Directly Related\n"
                                                f"Moderately Related\n"
                                                f"Slightly Related\n"
                                                f"Not Related"
                                            )
                                            
                                            # Use existing AI client logic
                                            client = anthropic.Anthropic()
                                            message = client.messages.create(
                                                model="claude-opus-4-6",
                                                max_tokens=20,
                                                messages=[{"role": "user", "content": prompt}]
                                            )
                                            result = message.content[0].text.strip()
                                            
                                            # Validate AI response
                                            valid = ["Directly Related", "Moderately Related", "Slightly Related", "Not Related"]
                                            if result not in valid:
                                                # Fuzzy match
                                                result_lower = result.lower()
                                                if "directly" in result_lower:
                                                    relevance = "Directly Related"
                                                elif "moderate" in result_lower:
                                                    relevance = "Moderately Related"
                                                elif "slightly" in result_lower:
                                                    relevance = "Slightly Related"
                                                else:
                                                    relevance = "Not Related"
                                            else:
                                                relevance = result
                                                
                                        except Exception as e:
                                            print(f"AI scan error: {e}")
                                            relevance = "Not Related"  # Fallback
                                    
                                    cursor.execute("""
                                        INSERT INTO alumni_employment
                                        (alumni_id, employment_status, employment_sector, job_title, degree_relevance_to_work, added_by, date_added)
                                        VALUES (%s,%s,%s,%s,%s,%s,%s)
                                    """, (alumni_id, employment_status, sector or None, job_title or None, relevance or None,
                                          "alumni_self_service", date.today()))
                            else:
                                # Create empty employment record
                                cursor.execute("""
                                    INSERT INTO alumni_employment
                                    (alumni_id, employment_status, employment_sector, job_title, degree_relevance_to_work, added_by, date_added)
                                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                                """, (alumni_id, employment_status, None, None, None, "alumni_self_service", date.today()))
                    
                    db.commit()
                    log_activity("Created own alumni profile")
                    success = "Your alumni profile has been created successfully!"
                    
                    # Refresh data to ensure latest information is displayed
                    cursor.execute("SELECT * FROM alumni_table WHERE alumni_id=%s", (alumni_id,))
                    alumni = cursor.fetchone()
                    alumni['has_record'] = True
                    
                    cursor.execute("SELECT * FROM alumni_degree WHERE alumni_id=%s LIMIT 1", (alumni_id,))
                    educ = cursor.fetchone()
                    
                    cursor.execute("SELECT * FROM alumni_employment WHERE alumni_id=%s ORDER BY employment_id ASC", (alumni_id,))
                    employ = cursor.fetchall()
                    
                else:
                    # ---------- ALUMNI UPDATING RECORDS WITH APPROVED REQUEST ----------
                    # Allow updates when they have an approved request
                    if not can_edit:
                        error = "Your update request is still pending. Please wait for admin approval."
                    else:
                        # ---------- VALIDATION ----------
                        # Student Number validation
                        stud_num = request.form.get("stud_num", "").strip()
                        if not stud_num:
                            error = "Student number is required."
                        else:
                            # Remove TAL- prefix if present and validate format
                            clean_stud_num = stud_num.replace("TAL-", "").replace(" ", "")
                            if not re.match(r"^\d{4}-\d{5}$", clean_stud_num):
                                error = "Student number must be in format YYYY-XXXXX (e.g., 2023-00238)."
                            else:
                                # Auto-prepend TAL- for database storage
                                stud_num = f"TAL-{clean_stud_num}"
                        if not error:
                            # Check for duplicate student number (excluding own record)
                            cursor.execute("SELECT alumni_id FROM alumni_table WHERE stud_num=%s AND alumni_id != %s", (stud_num, alumni.get('alumni_id')))
                            if cursor.fetchone():
                                error = "This student number already exists. Please use your unique student number."
                        
                        # Email validation
                        email = request.form.get("email", "").strip()
                        if not email:
                            error = "Email is required." if not error else error
                        elif not re.match(r"^[^@]+@[^@]+\.[^@]+$", email):
                            error = "Please enter a valid email address." if not error else error
                        
                        # Contact validation
                        contact = re.sub(r"\D", "", request.form.get("contact_num", ""))
                        if not contact or len(contact) != 11 or not contact.startswith("09"):
                            error = "Contact number must be 11 digits starting with 09." if not error else error
                        
                        # Address validation
                        address = request.form.get("address", "").strip()
                        if not address:
                            error = "Address is required." if not error else error
                        
                        # ---------- Photo ----------
                        old_photo = alumni.get("photo")
                        remove_flag = request.form.get("remove_photo", "0")
                        photo_file = request.files.get("photo")
                        new_photo = old_photo

                        if remove_flag == "1":
                            delete_photo(old_photo)
                            new_photo = None
                        elif photo_file and photo_file.filename:
                            saved = save_photo(photo_file)
                            if saved:
                                delete_photo(old_photo)
                                new_photo = saved
                        
                        # ---------- UPDATE ALUMNI RECORD ----------
                        if not error:
                            cursor.execute("""
                                UPDATE alumni_table 
                                SET stud_num = %s, last_name = %s, middle_name = %s, first_name = %s, 
                                    address = %s, email = %s, contact_num = %s, photo = %s,
                                    updated_by = %s, date_updated = %s
                                WHERE alumni_id = %s
                            """, (
                                stud_num,
                                request.form.get("last_name", "").strip(),
                                request.form.get("middle_name", "").strip(),
                                request.form.get("first_name", "").strip(),
                                address,
                                email,
                                contact,
                                new_photo,
                                "alumni_self_update",
                                date.today(),
                                alumni.get('alumni_id')
                            ))
                            
                            # ---------- DEGREE INFORMATION ----------
                            program = request.form.get("program", "").strip()
                            major = request.form.get("major", "").strip()
                            grad_date = request.form.get("graduation_date") or None
                            
                            if program:
                                # Check if degree record exists
                                cursor.execute("SELECT alumni_id FROM alumni_degree WHERE alumni_id=%s", (alumni.get('alumni_id'),))
                                degree_exists = cursor.fetchone()
                                
                                if degree_exists:
                                    # Update existing degree record
                                    cursor.execute("""
                                        UPDATE alumni_degree 
                                        SET program = %s, major = %s, graduation_date = %s,
                                            updated_by = %s, date_updated = %s
                                        WHERE alumni_id = %s
                                    """, (
                                        program, major, grad_date,
                                        "alumni_self_update", date.today(),
                                        alumni.get('alumni_id')
                                    ))
                                else:
                                    # Insert new degree record
                                    cursor.execute("""
                                        INSERT INTO alumni_degree
                                        (alumni_id, program, major, graduation_date, added_by, date_added)
                                        VALUES (%s,%s,%s,%s,%s,%s)
                                    """, (
                                        alumni.get('alumni_id'), program, major, grad_date,
                                        "alumni_self_update", date.today()
                                    ))
                            
                            # ---------- EMPLOYMENT INFORMATION ----------
                            employment_status = request.form.get("employment_status", "").strip()
                            if employment_status:
                                # First delete existing employment records for this alumni
                                cursor.execute("DELETE FROM alumni_employment WHERE alumni_id=%s", (alumni.get('alumni_id'),))
                                
                                if employment_status == "Unemployed":
                                    # Insert single unemployed record
                                    cursor.execute("""
                                        INSERT INTO alumni_employment
                                        (alumni_id, employment_status, employment_sector, job_title, degree_relevance_to_work, updated_by, date_updated)
                                        VALUES (%s,%s,%s,%s,%s,%s,%s)
                                    """, (
                                        alumni.get('alumni_id'), "Unemployed", None, None, None,
                                        "alumni_self_update", date.today()
                                    ))
                                else:
                                    # Handle employed/self-employed with multiple job details
                                    job_titles = request.form.getlist("job_title[]")
                                    sectors = request.form.getlist("employment_sector[]")
                                    relevances = request.form.getlist("degree_relevance_to_work[]")
                                    
                                    if job_titles:
                                        for i, job_title in enumerate(job_titles):
                                            if job_title.strip():  # Only process non-empty job titles
                                                sector = sectors[i] if i < len(sectors) else ""
                                                relevance = relevances[i] if i < len(relevances) else ""
                                                
                                                # Handle "Let AI Decide" option
                                                if relevance == "Let AI decide" or relevance == "Let AI Decide":
                                                    # Call AI scan relevance function
                                                    try:
                                                        program = request.form.get("program", "")
                                                        major = request.form.get("major", "")
                                                        
                                                        # Prepare AI scan request
                                                        ai_data = {
                                                            "job_title": job_title.strip(),
                                                            "program": program,
                                                            "major": major
                                                        }
                                                        
                                                        # Import the scan function logic here
                                                        prompt = (
                                                            f"A college graduate completed a {program} degree "
                                                            f"with a major in {major}. "
                                                            f"Their current job title is: '{job_title.strip()}'.\n\n"
                                                            f"Based on this, classify how relevant their degree is to their work. "
                                                            f"Reply with EXACTLY one of these four options and nothing else:\n"
                                                            f"Directly Related\n"
                                                            f"Moderately Related\n"
                                                            f"Slightly Related\n"
                                                            f"Not Related"
                                                        )
                                                        
                                                        # Use existing AI client logic
                                                        client = anthropic.Anthropic()
                                                        message = client.messages.create(
                                                            model="claude-opus-4-6",
                                                            max_tokens=20,
                                                            messages=[{"role": "user", "content": prompt}]
                                                        )
                                                        result = message.content[0].text.strip()
                                                        
                                                        # Validate AI response
                                                        valid = ["Directly Related", "Moderately Related", "Slightly Related", "Not Related"]
                                                        if result not in valid:
                                                            # Fuzzy match
                                                            result_lower = result.lower()
                                                            if "directly" in result_lower:
                                                                relevance = "Directly Related"
                                                            elif "moderate" in result_lower:
                                                                relevance = "Moderately Related"
                                                            elif "slightly" in result_lower:
                                                                relevance = "Slightly Related"
                                                            else:
                                                                relevance = "Not Related"
                                                        else:
                                                            relevance = result
                                                            
                                                    except Exception as e:
                                                        print(f"AI scan error: {e}")
                                                        relevance = "Not Related"  # Fallback
                                                
                                                cursor.execute("""
                                                    INSERT INTO alumni_employment
                                                    (alumni_id, employment_status, employment_sector, job_title, degree_relevance_to_work, updated_by, date_updated)
                                                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                                                """, (
                                                    alumni.get('alumni_id'), employment_status, sector.strip() if sector else None, 
                                                    job_title.strip(), relevance.strip() if relevance else None,
                                                    "alumni_self_update", date.today()
                                                ))
                                    else:
                                        # Create single employment record with no job details
                                        cursor.execute("""
                                            INSERT INTO alumni_employment
                                            (alumni_id, employment_status, employment_sector, job_title, degree_relevance_to_work, updated_by, date_updated)
                                            VALUES (%s,%s,%s,%s,%s,%s,%s)
                                        """, (
                                            alumni.get('alumni_id'), employment_status, None, None, None,
                                            "alumni_self_update", date.today()
                                        ))
                            
                            # Mark update request as completed
                            request_alumni_id = alumni.get('alumni_id')
                            cursor.execute("""
                                UPDATE alumni_update_requests 
                                SET status = 'completed', resolved_at = %s, resolved_by = %s
                                WHERE alumni_id = %s AND status = 'approved'
                            """, (
                                datetime.now(),
                                session.get("username"),
                                request_alumni_id
                            ))
                            
                            db.commit()
                            success = "Record Successfully Updated"
                            log_activity("Updated own alumni record")
                            
                            # Refresh data to ensure latest information is displayed
                            cursor.execute("SELECT * FROM alumni_table WHERE alumni_id=%s", (alumni.get('alumni_id'),))
                            alumni = cursor.fetchone()
                            alumni['has_record'] = True
                            
                            cursor.execute("SELECT * FROM alumni_degree WHERE alumni_id=%s LIMIT 1", (alumni.get('alumni_id'),))
                            educ = cursor.fetchone()
                            
                            cursor.execute("SELECT * FROM alumni_employment WHERE alumni_id=%s ORDER BY employment_id ASC", (alumni.get('alumni_id'),))
                            employ = cursor.fetchall()
                            
                            # Update session with fresh data if needed
                            if alumni:
                                # Properly format the full name
                                full_name = f"{alumni.get('last_name', '')}, {alumni.get('first_name', '')}"
                                if alumni.get('middle_name') and alumni.get('middle_name').strip():
                                    full_name += f" {alumni.get('middle_name')[0]}."
                                session['alumni_fullname'] = full_name

        except Exception as e:
            db.rollback()
            error = str(e)

    # Check if this is an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.args.get('ajax'):
        # Close database connection before returning JSON
        cursor.close()
        db.close()
        # Return JSON response for AJAX requests
        return jsonify({
            'success': not error,
            'error': error
        })
    
    # Return HTML template for regular requests
    # Close database connection before rendering template
    cursor.close()
    db.close()
    
    return render_template(
        "my_profile.html",
        alumni=alumni,
        degree=educ,
        employment=employ,
        programs=programs,
        error=error,
        success=success,
        can_edit=can_edit,
        user_type="ALUMNI",
        now=datetime.now()
    )









# ================= Alumni Notifications (ADMIN ONLY) =================
@app.route("/alumni-notif")
@admin_required
def alumni_notif():
    filter_status = request.args.get("status", "pending")

    db = get_db()
    cursor = db.cursor()

    if filter_status == "all":
        where = ""
        params = []
    else:
        where = "WHERE r.status = %s"
        params = [filter_status]

    cursor.execute(f"""
        SELECT
            r.id,
            r.alumni_id,
            r.reason,
            r.status,
            r.admin_note,
            r.requested_at,
            r.resolved_at,
            r.resolved_by,
            CONCAT(a.first_name, ' ', a.last_name) AS full_name,
            a.stud_num,
            a.photo
        FROM alumni_update_requests r
        LEFT JOIN alumni_table a ON r.alumni_id = a.alumni_id
        {where}
        ORDER BY r.requested_at DESC
    """, params)

    notifications = cursor.fetchall()

    cursor.execute("SELECT COUNT(*) as c FROM alumni_update_requests WHERE status = 'pending'")
    result = cursor.fetchone()
    pending_count = result['c'] if result else 0

    cursor.execute("SELECT COUNT(*) as c FROM alumni_update_requests WHERE status = 'approved'")
    result = cursor.fetchone()
    approved_count = result['c'] if result else 0

    cursor.execute("SELECT COUNT(*) as c FROM alumni_update_requests WHERE status = 'rejected'")
    result = cursor.fetchone()
    rejected_count = result['c'] if result else 0

    cursor.close()
    db.close()

    return render_template(
        "alumni_notif.html",
        notifications=notifications,
        pending_count=pending_count,
        approved_count=approved_count,
        rejected_count=rejected_count,
        filter_status=filter_status,
        user_type=session.get("user_type")
    )


# ================= API: Pending notif count (for sidebar polling) =================
@app.route("/api/pending-notif-count")
@login_required
def pending_notif_count():
    if session.get("user_type") != "Admin":
        return jsonify({"count": 0})

    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT COUNT(*) as count FROM alumni_update_requests WHERE status = 'pending'")
    result = cursor.fetchone()
    count = result['count'] if result else 0
    cursor.close()
    db.close()

    return jsonify({"count": count})


# ================= API: Alumni sends update request =================
@app.route("/request-update", methods=["POST"])
@login_required
def request_update():
    if session.get("user_type") != "Alumni":
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    data = request.get_json(force=True)
    reason = (data.get("reason") or "").strip() or "No reason provided."

    db = get_db()
    cursor = db.cursor()

    try:
        # Get alumni info from alumni_table
        cursor.execute("SELECT * FROM alumni_table WHERE email = %s", (session.get("username"),))
        alumni_record = cursor.fetchone()

        if not alumni_record:
            return jsonify({"success": False, "error": "Account not found."})
        else:
            alumni_id = alumni_record['alumni_id']
            print(f"DEBUG: Found existing alumni record with ID: {alumni_id}")

        # Prevent duplicate pending requests
        cursor.execute(
            "SELECT id FROM alumni_update_requests WHERE alumni_id = %s AND status = 'pending'",
            (alumni_id,)
        )
        existing = cursor.fetchone()

        if existing:
            return jsonify({"success": False, "error": "You already have a pending request. Please wait for Admin approval."})

        cursor.execute("""
            INSERT INTO alumni_update_requests (alumni_id, reason, status, requested_at)
            VALUES (%s, %s, 'pending', %s)
        """, (alumni_id, reason, datetime.now()))

        db.commit()
        log_activity("Sent update request")

        return jsonify({"success": True})

    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "error": str(e)})

    finally:
        cursor.close()
        db.close()


# ================= API: Admin approves or rejects request =================
@app.route("/api/resolve-update-request", methods=["POST"])
@admin_required
def resolve_update_request():
    data = request.get_json(force=True)
    req_id = data.get("id")
    action = data.get("action")  # 'approved' or 'rejected'
    note = (data.get("note") or "").strip()

    if action not in ("approved", "rejected"):
        return jsonify({"success": False, "error": "Invalid action."})

    db = get_db()
    cursor = db.cursor()

    try:
        # Get the request details including update_data
        cursor.execute("SELECT alumni_id, update_data FROM alumni_update_requests WHERE id = %s", (req_id,))
        request_data = cursor.fetchone()
        
        if not request_data:
            return jsonify({"success": False, "error": "Request not found."})
        
        alumni_id = request_data['alumni_id']  # FIXED: Use dictionary access
        
        # If approved, apply the updates to alumni records
        if action == 'approved':
            import json
            update_data = request_data.get('update_data')
            if update_data:
                try:
                    updates = json.loads(update_data)
                    
                    # Update alumni_table basic info
                    if 'alumni_table' in updates:
                        basic_updates = updates['alumni_table']
                        update_fields = []
                        update_values = []
                        
                        for field, value in basic_updates.items():
                            if field in ['last_name', 'first_name', 'middle_name', 'address', 'email', 'contact_num']:
                                update_fields.append(f"{field} = %s")
                                update_values.append(value)
                        
                        if update_fields:
                            update_fields.append("updated_by = %s")
                            update_fields.append("date_updated = %s")
                            update_values.append(session.get("username"))
                            update_values.append(datetime.now())
                            update_values.append(alumni_id)
                            
                            cursor.execute(f"""
                                UPDATE alumni_table 
                                SET {', '.join(update_fields)}
                                WHERE alumni_id = %s
                            """, update_values)
                    
                    # Update alumni_degree if present
                    if 'alumni_degree' in updates:
                        degree_updates = updates['alumni_degree']
                        update_fields = []
                        update_values = []
                        
                        for field, value in degree_updates.items():
                            if field in ['program', 'major', 'graduation_date']:
                                update_fields.append(f"{field} = %s")
                                update_values.append(value)
                        
                        if update_fields:
                            update_fields.append("updated_by = %s")
                            update_fields.append("date_updated = %s")
                            update_values.append(session.get("username"))
                            update_values.append(datetime.now())
                            update_values.append(alumni_id)
                            
                            cursor.execute(f"""
                                UPDATE alumni_degree 
                                SET {', '.join(update_fields)}
                                WHERE alumni_id = %s
                            """, update_values)
                    
                    # Update alumni_employment if present
                    if 'alumni_employment' in updates:
                        emp_updates = updates['alumni_employment']
                        update_fields = []
                        update_values = []
                        
                        for field, value in emp_updates.items():
                            if field in ['employment_status', 'job_title', 'employment_sector', 'degree_relevance_to_work']:
                                update_fields.append(f"{field} = %s")
                                update_values.append(value)
                        
                        if update_fields:
                            update_fields.append("updated_by = %s")
                            update_fields.append("date_updated = %s")
                            update_values.append(session.get("username"))
                            update_values.append(datetime.now())
                            update_values.append(alumni_id)
                            
                            cursor.execute(f"""
                                UPDATE alumni_employment 
                                SET {', '.join(update_fields)}
                                WHERE alumni_id = %s
                            """, update_values)
                            
                except json.JSONDecodeError:
                    print(f"Warning: Could not parse update_data for request {req_id}")
        
        # Update the request status
        cursor.execute("""
            UPDATE alumni_update_requests
            SET status = %s,
                admin_note = %s,
                resolved_at = %s,
                resolved_by = %s
            WHERE id = %s
        """, (action, note or None, datetime.now(), session.get("username"), req_id))

        db.commit()
        log_activity(f"{'Approved' if action == 'approved' else 'Rejected'} alumni update request #{req_id}")

        return jsonify({"success": True})

    except Exception as e:
        db.rollback()
        print(f"Error in resolve_update_request: {str(e)}")  # Debug output
        return jsonify({"success": False, "error": str(e)})

    finally:
        cursor.close()
        db.close()


# ================= API: Alumni checks if their request is approved =================
@app.route("/api/my-update-status")
@login_required
def my_update_status():
    if session.get("user_type") != "Alumni":
        return jsonify({"error": "Unauthorized"}), 403

    db = get_db()
    cursor = db.cursor()

    # Get alumni info from alumni_table
    cursor.execute("SELECT alumni_id, email FROM alumni_table WHERE email = %s LIMIT 1", (session.get("username"),))
    alumni_record = cursor.fetchone()
    
    if not alumni_record:
        cursor.close()
        db.close()
        return jsonify({"status": "none"})
    
    # Use alumni_table ID
    alumni_id = alumni_record['alumni_id']
    
    cursor.execute(""" 
        SELECT status, admin_note, resolved_at
        FROM alumni_update_requests
        WHERE alumni_id = %s
        ORDER BY requested_at DESC
        LIMIT 1
    """, (alumni_id,))

    result = cursor.fetchone()
    cursor.close()
    db.close()

    if result:
        return jsonify({
            "status": result["status"],
            "admin_note": result["admin_note"],
            "resolved_at": result["resolved_at"].strftime("%Y-%m-%d %H:%M") if result["resolved_at"] else None
        })

    return jsonify({"status": "none"})


# ================= API: Coordinator checks their announcement request status =================
@app.route("/api/my-announcement-status")
@login_required
def my_announcement_status():
    if session.get("user_type") != "ALUMNI COORDINATOR":
        return jsonify({"error": "Unauthorized"}), 403

    db = get_db()
    cursor = db.cursor()

    cursor.execute("""
        SELECT id, subject, message, recipient_emails, status, created_at, approved_at, approved_by, admin_note
        FROM announcement_requests
        WHERE coordinator_id = %s
        ORDER BY created_at DESC
        LIMIT 10
    """, (session.get("username"),))

    requests = cursor.fetchall()
    cursor.close()
    db.close()

    # Format dates for JSON response
    for req in requests:
        if req['created_at']:
            req['created_at'] = req['created_at'].strftime("%Y-%m-%d %H:%M")
        if req['approved_at']:
            req['approved_at'] = req['approved_at'].strftime("%Y-%m-%d %H:%M")

    return jsonify({"requests": requests})


# ================= REPORTS API ROUTES =================
@app.route("/api/report-statistics")
@login_required
def report_statistics():
    db = get_db()
    cursor = db.cursor()
    
    try:
        # Get basic statistics with proper joins
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT a.alumni_id) as total_alumni,
                SUM(CASE WHEN e.employment_status = 'Employed' THEN 1 ELSE 0 END) as employed_alumni,
                SUM(CASE WHEN e.employment_status = 'Unemployed' THEN 1 ELSE 0 END) as unemployed_alumni,
                SUM(CASE WHEN e.employment_status = 'Self-Employed' THEN 1 ELSE 0 END) as self_employed,
                SUM(CASE WHEN e.degree_relevance_to_work = 'Directly Related' THEN 1 ELSE 0 END) as relevant_work
            FROM alumni_table a
            LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
            WHERE a.email IS NOT NULL AND TRIM(a.email) != ''
        """)
        
        stats = cursor.fetchone()
        
        # Calculate percentages safely
        total = stats['total_alumni'] or 0
        if total > 0:
            stats['employed_percentage'] = round((stats['employed_alumni'] or 0) / total * 100, 1)
            stats['unemployed_percentage'] = round((stats['unemployed_alumni'] or 0) / total * 100, 1)
            stats['self_employed_percentage'] = round((stats['self_employed'] or 0) / total * 100, 1)
            stats['relevant_work_percentage'] = round((stats['relevant_work'] or 0) / total * 100, 1)
        else:
            stats['employed_percentage'] = 0
            stats['unemployed_percentage'] = 0
            stats['self_employed_percentage'] = 0
            stats['relevant_work_percentage'] = 0
        
        cursor.close()
        db.close()
        return jsonify(stats)
        
    except Exception as e:
        print("Report statistics error:", e)
        cursor.close()
        db.close()
        return jsonify({
            'total_alumni': 0,
            'employed_alumni': 0,
            'unemployed_alumni': 0,
            'self_employed': 0,
            'relevant_work': 0,
            'employed_percentage': 0,
            'unemployed_percentage': 0,
            'self_employed_percentage': 0,
            'relevant_work_percentage': 0
        })


@app.route("/api/generate-report/<report_type>")
@login_required
def generate_report(report_type):
    db = get_db()
    cursor = db.cursor()
    
    try:
        if report_type == "summary":
            # Summary report data
            cursor.execute("""
                SELECT 
                    COUNT(DISTINCT a.alumni_id) as total_alumni,
                    SUM(CASE WHEN e.employment_status = 'Employed' THEN 1 ELSE 0 END) as employed_alumni,
                    SUM(CASE WHEN e.employment_status = 'Unemployed' THEN 1 ELSE 0 END) as unemployed_alumni,
                    SUM(CASE WHEN e.employment_status = 'Self-Employed' THEN 1 ELSE 0 END) as self_employed,
                    SUM(CASE WHEN e.degree_relevance_to_work = 'Directly Related' THEN 1 ELSE 0 END) as relevant_work
                FROM alumni_table a
                LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
                WHERE a.email IS NOT NULL AND TRIM(a.email) != ''
            """)
            data = cursor.fetchone()
            
            total = data['total_alumni'] or 0
            if total > 0:
                data['employed_percentage'] = round((data['employed_alumni'] or 0) / total * 100, 1)
                data['unemployed_percentage'] = round((data['unemployed_alumni'] or 0) / total * 100, 1)
                data['self_employed_percentage'] = round((data['self_employed'] or 0) / total * 100, 1)
                data['relevant_work_percentage'] = round((data['relevant_work'] or 0) / total * 100, 1)
            else:
                data['employed_percentage'] = 0
                data['unemployed_percentage'] = 0
                data['self_employed_percentage'] = 0
                data['relevant_work_percentage'] = 0
                
        elif report_type == "graduates-year":
            # Graduates per year report
            cursor.execute("""
                SELECT 
                    YEAR(d.graduation_date) as year,
                    COUNT(DISTINCT a.alumni_id) as total,
                    SUM(CASE WHEN d.program = 'BSIT' THEN 1 ELSE 0 END) as bsit,
                    SUM(CASE WHEN d.program = 'BSBA' THEN 1 ELSE 0 END) as bsba,
                    SUM(CASE WHEN d.program = 'BEED' THEN 1 ELSE 0 END) as beed
                FROM alumni_table a
                LEFT JOIN alumni_degree d ON a.alumni_id = d.alumni_id
                WHERE d.graduation_date IS NOT NULL
                GROUP BY YEAR(d.graduation_date)
                ORDER BY year DESC
                LIMIT 10
            """)
            data = {"graduates_by_year": cursor.fetchall()}
            
        elif report_type == "employment-status":
            # Employment status report
            cursor.execute("""
                SELECT 
                    COALESCE(e.employment_status, 'Unknown') as status,
                    COUNT(DISTINCT a.alumni_id) as count,
                    ROUND(COUNT(DISTINCT a.alumni_id) * 100.0 / (SELECT COUNT(DISTINCT alumni_id) FROM alumni_table WHERE email IS NOT NULL AND TRIM(email) != ''), 1) as percentage
                FROM alumni_table a
                LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
                WHERE a.email IS NOT NULL AND TRIM(a.email) != ''
                GROUP BY e.employment_status
                ORDER BY count DESC
            """)
            data = {"employment_status": cursor.fetchall()}
            
        elif report_type == "program-stats":
            # Program statistics report
            cursor.execute("""
                SELECT 
                    d.program,
                    COUNT(DISTINCT a.alumni_id) as count,
                    ROUND(COUNT(DISTINCT a.alumni_id) * 100.0 / (SELECT COUNT(DISTINCT alumni_id) FROM alumni_table WHERE email IS NOT NULL AND TRIM(email) != ''), 1) as percentage,
                    SUM(CASE WHEN e.employment_status = 'Employed' THEN 1 ELSE 0 END) as employed_count,
                    ROUND(SUM(CASE WHEN e.employment_status = 'Employed' THEN 1 ELSE 0 END) * 100.0 / COUNT(DISTINCT a.alumni_id), 1) as employment_rate
                FROM alumni_table a
                LEFT JOIN alumni_degree d ON a.alumni_id = d.alumni_id
                LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
                WHERE a.email IS NOT NULL AND TRIM(a.email) != '' AND d.program IS NOT NULL
                GROUP BY d.program
                ORDER BY count DESC
            """)
            data = {"program_stats": cursor.fetchall()}
            
        elif report_type == "sector-stats":
            # Employment sector statistics
            cursor.execute("""
                SELECT 
                    COALESCE(e.employment_sector, 'Unknown') as sector,
                    COUNT(DISTINCT a.alumni_id) as count,
                    ROUND(COUNT(DISTINCT a.alumni_id) * 100.0 / (SELECT COUNT(DISTINCT alumni_id) FROM alumni_employment WHERE employment_sector IS NOT NULL), 1) as percentage
                FROM alumni_table a
                LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
                WHERE e.employment_sector IS NOT NULL
                GROUP BY e.employment_sector
                ORDER BY count DESC
            """)
            data = {"sector_stats": cursor.fetchall()}
            
        elif report_type == "relevance-stats":
            # Work relevance statistics
            cursor.execute("""
                SELECT 
                    COALESCE(e.degree_relevance_to_work, 'Unknown') as relevance,
                    COUNT(DISTINCT a.alumni_id) as count,
                    ROUND(COUNT(DISTINCT a.alumni_id) * 100.0 / (SELECT COUNT(DISTINCT alumni_id) FROM alumni_employment WHERE degree_relevance_to_work IS NOT NULL), 1) as percentage
                FROM alumni_table a
                LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
                WHERE e.degree_relevance_to_work IS NOT NULL
                GROUP BY e.degree_relevance_to_work
                ORDER BY count DESC
            """)
            data = {"relevance_stats": cursor.fetchall()}
            
        elif report_type == "tracking":
            # Employment tracking report
            cursor.execute("""
                SELECT 
                    CONCAT(a.first_name, ' ', a.last_name) as name,
                    d.program,
                    YEAR(d.graduation_date) as graduation_year,
                    COALESCE(e.employment_status, 'Unknown') as employment_status,
                    e.job_title,
                    COALESCE(e.employment_sector, 'Unknown') as employment_sector
                FROM alumni_table a
                LEFT JOIN alumni_degree d ON a.alumni_id = d.alumni_id
                LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
                WHERE a.email IS NOT NULL AND TRIM(a.email) != ''
                ORDER BY d.graduation_date DESC
                LIMIT 50
            """)
            data = {"tracking_data": cursor.fetchall()}
            
        elif report_type == "recent":
            # Recently added alumni report
            cursor.execute("""
                SELECT 
                    CONCAT(a.first_name, ' ', a.last_name) as name,
                    d.program,
                    d.major,
                    DATE(a.date_added) as date_added,
                    a.added_by
                FROM alumni_table a
                LEFT JOIN alumni_degree d ON a.alumni_id = d.alumni_id
                WHERE a.date_added >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
                ORDER BY a.date_added DESC
            """)
            data = {"recent_alumni": cursor.fetchall()}
            
        else:
            return jsonify({"error": "Invalid report type"}), 400
        
        cursor.close()
        db.close()
        return jsonify(data)
        
    except Exception as e:
        print(f"Generate report error for {report_type}:", e)
        cursor.close()
        db.close()
        return jsonify({"error": str(e)}), 500


@app.route("/api/export-pdf/<report_type>")
@login_required
def export_pdf(report_type):
    try:
        # Generate PDF report
        pdf_buffer = generate_pdf_report(report_type)
        
        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=f"{report_type}_report_{datetime.now().strftime('%Y%m%d')}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/export-excel/<report_type>")
@login_required
def export_excel(report_type):
    try:
        # Generate Excel report
        excel_buffer = generate_excel_report(report_type)
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=f"{report_type}_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def generate_pdf_report(report_type):
    # Get report data
    db = get_db()
    cursor = db.cursor()
    
    try:
        # Fetch data based on report type
        if report_type == "summary":
            cursor.execute("""
                SELECT 
                    COUNT(DISTINCT a.alumni_id) as total_alumni,
                    SUM(CASE WHEN e.employment_status = 'Employed' THEN 1 ELSE 0 END) as employed_alumni,
                    SUM(CASE WHEN e.employment_status = 'Unemployed' THEN 1 ELSE 0 END) as unemployed_alumni,
                    SUM(CASE WHEN e.employment_status = 'Self-Employed' THEN 1 ELSE 0 END) as self_employed,
                    SUM(CASE WHEN e.degree_relevance_to_work = 'Directly Related' THEN 1 ELSE 0 END) as relevant_work
                FROM alumni_table a
                LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
                WHERE a.email IS NOT NULL AND TRIM(a.email) != ''
            """)
            data = cursor.fetchone()
            
            total = data['total_alumni'] or 0
            if total > 0:
                data['employed_percentage'] = round((data['employed_alumni'] or 0) / total * 100, 1)
                data['unemployed_percentage'] = round((data['unemployed_alumni'] or 0) / total * 100, 1)
                data['self_employed_percentage'] = round((data['self_employed'] or 0) / total * 100, 1)
                data['relevant_work_percentage'] = round((data['relevant_work'] or 0) / total * 100, 1)
            else:
                data['employed_percentage'] = 0
                data['unemployed_percentage'] = 0
                data['self_employed_percentage'] = 0
                data['relevant_work_percentage'] = 0
        elif report_type == "employment-status":
            cursor.execute("""
                SELECT 
                    COALESCE(e.employment_status, 'Unknown') as status,
                    COUNT(DISTINCT a.alumni_id) as count,
                    ROUND(COUNT(DISTINCT a.alumni_id) * 100.0 / (SELECT COUNT(DISTINCT alumni_id) FROM alumni_table WHERE email IS NOT NULL AND TRIM(email) != ''), 1) as percentage
                FROM alumni_table a
                LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
                WHERE a.email IS NOT NULL AND TRIM(a.email) != ''
                GROUP BY e.employment_status
                ORDER BY count DESC
            """)
            data = {"employment_status": cursor.fetchall()}
        elif report_type == "program-stats":
            cursor.execute("""
                SELECT 
                    d.program,
                    COUNT(DISTINCT a.alumni_id) as count,
                    ROUND(COUNT(DISTINCT a.alumni_id) * 100.0 / (SELECT COUNT(DISTINCT alumni_id) FROM alumni_table WHERE email IS NOT NULL AND TRIM(email) != ''), 1) as percentage,
                    SUM(CASE WHEN e.employment_status = 'Employed' THEN 1 ELSE 0 END) as employed_count,
                    ROUND(SUM(CASE WHEN e.employment_status = 'Employed' THEN 1 ELSE 0 END) * 100.0 / COUNT(DISTINCT a.alumni_id), 1) as employment_rate
                FROM alumni_table a
                LEFT JOIN alumni_degree d ON a.alumni_id = d.alumni_id
                LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
                WHERE a.email IS NOT NULL AND TRIM(a.email) != '' AND d.program IS NOT NULL
                GROUP BY d.program
                ORDER BY count DESC
            """)
            data = {"program_stats": cursor.fetchall()}
        else:
            # Default to summary for other types
            cursor.execute("""
                SELECT 
                    COUNT(DISTINCT a.alumni_id) as total_alumni,
                    SUM(CASE WHEN e.employment_status = 'Employed' THEN 1 ELSE 0 END) as employed_alumni,
                    SUM(CASE WHEN e.employment_status = 'Unemployed' THEN 1 ELSE 0 END) as unemployed_alumni,
                    SUM(CASE WHEN e.employment_status = 'Self-Employed' THEN 1 ELSE 0 END) as self_employed,
                    SUM(CASE WHEN e.degree_relevance_to_work = 'Directly Related' THEN 1 ELSE 0 END) as relevant_work
                FROM alumni_table a
                LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
                WHERE a.email IS NOT NULL AND TRIM(a.email) != ''
            """)
            data = cursor.fetchone()
            
            total = data['total_alumni'] or 0
            if total > 0:
                data['employed_percentage'] = round((data['employed_alumni'] or 0) / total * 100, 1)
                data['unemployed_percentage'] = round((data['unemployed_alumni'] or 0) / total * 100, 1)
                data['self_employed_percentage'] = round((data['self_employed'] or 0) / total * 100, 1)
                data['relevant_work_percentage'] = round((data['relevant_work'] or 0) / total * 100, 1)
            else:
                data['employed_percentage'] = 0
                data['unemployed_percentage'] = 0
                data['self_employed_percentage'] = 0
                data['relevant_work_percentage'] = 0
        
        cursor.close()
        db.close()
    except Exception as e:
        cursor.close()
        db.close()
        raise e
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=LETTER, rightMargin=50, leftMargin=50, topMargin=60, bottomMargin=60)
    elements = []
    
    # Add title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1
    )
    
    # Add logo
    try:
        logo = Image("static/images/PNG (transparent background).png", 50, 50)
    except:
        logo = None
    
    if logo:
        header_tbl = Table([[logo, Paragraph(
            "Nueva Ecija University of Science and Technology<br/>" +
            "Talavera Academic Extension Campus<br/><br/>" +
            f"<b>{report_type.replace('-', ' ').title()} Report</b>",
            title_style
        )]], colWidths=[60, 440])
        header_tbl.setStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")])
        elements.append(header_tbl)
    else:
        elements.append(Paragraph(f"{report_type.replace('-', ' ').title()} Report", title_style))
    
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Add content based on report type
    if report_type == "summary" or report_type not in ["employment-status", "program-stats"]:
        add_summary_to_pdf(elements, data)
    elif report_type == "employment-status":
        add_table_to_pdf(elements, data, "employment_status", ["Employment Status", "Count", "Percentage"])
    elif report_type == "program-stats":
        add_table_to_pdf(elements, data, "program_stats", ["Program", "Total Alumni", "Percentage", "Employed", "Employment Rate"])
    
    doc.build(elements, onFirstPage=draw_layout, onLaterPages=draw_layout)
    buffer.seek(0)
    return buffer


def generate_excel_report(report_type):
    # Get report data
    db = get_db()
    cursor = db.cursor()
    
    try:
        # Fetch data based on report type
        if report_type == "summary":
            cursor.execute("""
                SELECT 
                    COUNT(DISTINCT a.alumni_id) as total_alumni,
                    SUM(CASE WHEN e.employment_status = 'Employed' THEN 1 ELSE 0 END) as employed_alumni,
                    SUM(CASE WHEN e.employment_status = 'Unemployed' THEN 1 ELSE 0 END) as unemployed_alumni,
                    SUM(CASE WHEN e.employment_status = 'Self-Employed' THEN 1 ELSE 0 END) as self_employed,
                    SUM(CASE WHEN e.degree_relevance_to_work = 'Directly Related' THEN 1 ELSE 0 END) as relevant_work
                FROM alumni_table a
                LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
                WHERE a.email IS NOT NULL AND TRIM(a.email) != ''
            """)
            data = cursor.fetchone()
            
            total = data['total_alumni'] or 0
            if total > 0:
                data['employed_percentage'] = round((data['employed_alumni'] or 0) / total * 100, 1)
                data['unemployed_percentage'] = round((data['unemployed_alumni'] or 0) / total * 100, 1)
                data['self_employed_percentage'] = round((data['self_employed'] or 0) / total * 100, 1)
                data['relevant_work_percentage'] = round((data['relevant_work'] or 0) / total * 100, 1)
            else:
                data['employed_percentage'] = 0
                data['unemployed_percentage'] = 0
                data['self_employed_percentage'] = 0
                data['relevant_work_percentage'] = 0
        elif report_type == "employment-status":
            cursor.execute("""
                SELECT 
                    COALESCE(e.employment_status, 'Unknown') as status,
                    COUNT(DISTINCT a.alumni_id) as count,
                    ROUND(COUNT(DISTINCT a.alumni_id) * 100.0 / (SELECT COUNT(DISTINCT alumni_id) FROM alumni_table WHERE email IS NOT NULL AND TRIM(email) != ''), 1) as percentage
                FROM alumni_table a
                LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
                WHERE a.email IS NOT NULL AND TRIM(a.email) != ''
                GROUP BY e.employment_status
                ORDER BY count DESC
            """)
            data = {"employment_status": cursor.fetchall()}
        elif report_type == "program-stats":
            cursor.execute("""
                SELECT 
                    d.program,
                    COUNT(DISTINCT a.alumni_id) as count,
                    ROUND(COUNT(DISTINCT a.alumni_id) * 100.0 / (SELECT COUNT(DISTINCT alumni_id) FROM alumni_table WHERE email IS NOT NULL AND TRIM(email) != ''), 1) as percentage,
                    SUM(CASE WHEN e.employment_status = 'Employed' THEN 1 ELSE 0 END) as employed_count,
                    ROUND(SUM(CASE WHEN e.employment_status = 'Employed' THEN 1 ELSE 0 END) * 100.0 / COUNT(DISTINCT a.alumni_id), 1) as employment_rate
                FROM alumni_table a
                LEFT JOIN alumni_degree d ON a.alumni_id = d.alumni_id
                LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
                WHERE a.email IS NOT NULL AND TRIM(a.email) != '' AND d.program IS NOT NULL
                GROUP BY d.program
                ORDER BY count DESC
            """)
            data = {"program_stats": cursor.fetchall()}
        else:
            # Default to summary for other types
            cursor.execute("""
                SELECT 
                    COUNT(DISTINCT a.alumni_id) as total_alumni,
                    SUM(CASE WHEN e.employment_status = 'Employed' THEN 1 ELSE 0 END) as employed_alumni,
                    SUM(CASE WHEN e.employment_status = 'Unemployed' THEN 1 ELSE 0 END) as unemployed_alumni,
                    SUM(CASE WHEN e.employment_status = 'Self-Employed' THEN 1 ELSE 0 END) as self_employed,
                    SUM(CASE WHEN e.degree_relevance_to_work = 'Directly Related' THEN 1 ELSE 0 END) as relevant_work
                FROM alumni_table a
                LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
                WHERE a.email IS NOT NULL AND TRIM(a.email) != ''
            """)
            data = cursor.fetchone()
            
            total = data['total_alumni'] or 0
            if total > 0:
                data['employed_percentage'] = round((data['employed_alumni'] or 0) / total * 100, 1)
                data['unemployed_percentage'] = round((data['unemployed_alumni'] or 0) / total * 100, 1)
                data['self_employed_percentage'] = round((data['self_employed'] or 0) / total * 100, 1)
                data['relevant_work_percentage'] = round((data['relevant_work'] or 0) / total * 100, 1)
            else:
                data['employed_percentage'] = 0
                data['unemployed_percentage'] = 0
                data['self_employed_percentage'] = 0
                data['relevant_work_percentage'] = 0
        
        cursor.close()
        db.close()
    except Exception as e:
        cursor.close()
        db.close()
        raise e
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{report_type} Report"
    
    # Add headers and data based on report type
    if report_type == "summary" or report_type not in ["employment-status", "program-stats"]:
        ws.append(["Metric", "Count", "Percentage"])
        ws.append(["Total Alumni", data.get('total_alumni', 0), "100%"])
        ws.append(["Employed Alumni", data.get('employed_alumni', 0), f"{data.get('employed_percentage', 0)}%"])
        ws.append(["Unemployed Alumni", data.get('unemployed_alumni', 0), f"{data.get('unemployed_percentage', 0)}%"])
        ws.append(["Self-Employed", data.get('self_employed', 0), f"{data.get('self_employed_percentage', 0)}%"])
        ws.append(["Degree-Relevant Work", data.get('relevant_work', 0), f"{data.get('relevant_work_percentage', 0)}%"])
    elif report_type == "employment-status":
        ws.append(["Employment Status", "Count", "Percentage"])
        for item in data.get('employment_status', []):
            ws.append([item['status'], item['count'], f"{item['percentage']}%"])
    elif report_type == "program-stats":
        ws.append(["Program", "Total Alumni", "Percentage", "Employed", "Employment Rate"])
        for item in data.get('program_stats', []):
            ws.append([item['program'], item['count'], f"{item['percentage']}%", item.get('employed_count', 0), f"{item.get('employment_rate', 0)}%"])
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def fetch_report_data(cursor, report_type):
    # Similar logic to generate_report function but returns data
    if report_type == "summary":
        cursor.execute("""
            SELECT 
                COUNT(*) as total_alumni,
                SUM(CASE WHEN e.employment_status = 'Employed' THEN 1 ELSE 0 END) as employed_alumni,
                SUM(CASE WHEN e.employment_status = 'Unemployed' THEN 1 ELSE 0 END) as unemployed_alumni,
                SUM(CASE WHEN e.employment_status = 'Self-Employed' THEN 1 ELSE 0 END) as self_employed,
                SUM(CASE WHEN e.degree_relevance_to_work = 'Directly Related' THEN 1 ELSE 0 END) as relevant_work
            FROM alumni_table a
            LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
            WHERE a.email IS NOT NULL AND TRIM(a.email) != ''
        """)
        data = cursor.fetchone()
        
        total = data['total_alumni'] or 0
        if total > 0:
            data['employed_percentage'] = round((data['employed_alumni'] or 0) / total * 100, 1)
            data['unemployed_percentage'] = round((data['unemployed_alumni'] or 0) / total * 100, 1)
            data['self_employed_percentage'] = round((data['self_employed'] or 0) / total * 100, 1)
            data['relevant_work_percentage'] = round((data['relevant_work'] or 0) / total * 100, 1)
        else:
            data['employed_percentage'] = 0
            data['unemployed_percentage'] = 0
            data['self_employed_percentage'] = 0
            data['relevant_work_percentage'] = 0
        return data
    
    elif report_type == "employment-status":
        cursor.execute("""
            SELECT 
                e.employment_status as status,
                COUNT(*) as count,
                ROUND(COUNT(*) * 100.0 / (SELECT COUNT(*) FROM alumni_table WHERE email IS NOT NULL AND TRIM(email) != ''), 1) as percentage
            FROM alumni_table a
            LEFT JOIN alumni_employment e ON a.alumni_id = e.alumni_id
            WHERE a.email IS NOT NULL AND TRIM(a.email) != ''
            GROUP BY e.employment_status
            ORDER BY count DESC
        """)
        return {"employment_status": cursor.fetchall()}
    
    # Add other report types as needed
    return {}


def add_summary_to_pdf(elements, data):
    styles = getSampleStyleSheet()
    
    # Create table data
    table_data = [['Metric', 'Count', 'Percentage']]
    table_data.append(['Total Alumni', str(data.get('total_alumni', 0)), '100%'])
    table_data.append(['Employed Alumni', str(data.get('employed_alumni', 0)), f"{data.get('employed_percentage', 0)}%"])
    table_data.append(['Unemployed Alumni', str(data.get('unemployed_alumni', 0)), f"{data.get('unemployed_percentage', 0)}%"])
    table_data.append(['Self-Employed', str(data.get('self_employed', 0)), f"{data.get('self_employed_percentage', 0)}%"])
    table_data.append(['Degree-Relevant Work', str(data.get('relevant_work', 0)), f"{data.get('relevant_work_percentage', 0)}%"])
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)


def add_table_to_pdf(elements, data, key, headers):
    if key not in data or not data[key]:
        return
    
    styles = getSampleStyleSheet()
    
    # Create table data
    table_data = [headers]
    for item in data[key]:
        row = []
        for header in headers:
            # Map header to data key (simplified)
            if 'Status' in header:
                row.append(item.get('status', ''))
            elif 'Count' in header:
                row.append(str(item.get('count', 0)))
            elif 'Percentage' in header:
                row.append(f"{item.get('percentage', 0)}%")
            elif 'Program' in header:
                row.append(item.get('program', ''))
            elif 'Sector' in header:
                row.append(item.get('sector', ''))
            else:
                row.append('')
        table_data.append(row)
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)


# ================= PROGRAM/MAJOR MANAGEMENT API ROUTES =================
@app.route("/api/add-program-major", methods=["POST"])
@login_required
@admin_required
def add_program_major():
    try:
        program_name = request.form.get("programName", "").strip()
        program_description = request.form.get("programDescription", "").strip()
        major_names = request.form.getlist("majorName[]")
        major_descriptions = request.form.getlist("majorDescription[]")
        
        if not program_name or not program_description:
            return jsonify({"success": False, "message": "Program name and program description are required"})
        
        if not major_names or not major_descriptions or len(major_names) != len(major_descriptions):
            return jsonify({"success": False, "message": "At least one major with description is required"})
        
        # Validate all majors have names and descriptions
        for i, (major_name, major_desc) in enumerate(zip(major_names, major_descriptions)):
            if not major_name.strip() or not major_desc.strip():
                return jsonify({"success": False, "message": f"Major {i+1} name and description are required"})
        
        db = get_db()
        cursor = db.cursor()
        
        # Check if program already exists
        cursor.execute("SELECT program_id FROM programs WHERE program_name = %s", (program_name,))
        if cursor.fetchone():
            cursor.close()
            db.close()
            return jsonify({"success": False, "message": "Program already exists"})
        
        # Create new program
        cursor.execute("""
            INSERT INTO programs (program_name, program_description, created_by, date_created)
            VALUES (%s, %s, %s, NOW())
        """, (program_name, program_description, session.get("username")))
        program_id = cursor.lastrowid
        
        # Insert all majors for this program
        majors_added = 0
        for major_name, major_description in zip(major_names, major_descriptions):
            major_name = major_name.strip()
            major_description = major_description.strip()
            
            # Check if major already exists for this program
            cursor.execute("""
                SELECT major_id FROM majors 
                WHERE major_name = %s AND program_id = %s
            """, (major_name, program_id))
            if cursor.fetchone():
                continue  # Skip duplicate majors
            
            cursor.execute("""
                INSERT INTO majors (major_name, program_id, major_description, created_by, date_created)
                VALUES (%s, %s, %s, %s, NOW())
            """, (major_name, program_id, major_description, session.get("username")))
            majors_added += 1
        
        if majors_added == 0:
            cursor.close()
            db.close()
            return jsonify({"success": False, "message": "No new majors were added (all duplicates)"})
        
        db.commit()
        cursor.close()
        db.close()
        
        return jsonify({
            "success": True, 
            "message": f"Program and {majors_added} major(s) added successfully"
        })
        
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/get-programs")
@login_required
def get_programs():
    try:
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute("""
            SELECT program_name as name, program_description as description
            FROM programs
            ORDER BY program_name
        """)
        
        programs = cursor.fetchall()
        cursor.close()
        db.close()
        
        return jsonify({"programs": programs})
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/get-majors/<program_name>")
@login_required
def get_majors(program_name):
    try:
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute("""
            SELECT m.major_name as name, m.major_description as description
            FROM majors m
            JOIN programs p ON m.program_id = p.program_id
            WHERE p.program_name = %s
            ORDER BY m.major_name
        """, (program_name,))
        
        majors = cursor.fetchall()
        cursor.close()
        db.close()
        
        return jsonify({"majors": majors})
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 8080))
    # Production-safe: disable debug in production
    debug_mode = os.environ.get("FLASK_ENV") != "production"
    app.run(host="0.0.0.0", port=port, debug=debug_mode)